"""
GNC-쿠팡 자동 매칭 시스템
🔄 리팩토링: coupang_manager 모듈 사용
"""

import pandas as pd
from openpyxl import load_workbook
import time
import sys
import os
from typing import List, Optional, Dict
from dataclasses import dataclass, asdict
from datetime import datetime

# 프로젝트 루트
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
sys.path.insert(0, project_root)

# coupang_manager 사용
from coupang_manager import CoupangBrowser
from gnc_crawler import GNCCrawler, GNCProduct
from coupang_crawler import CoupangCrawler, CoupangProduct
from gemini_matcher import ImageMatcher, CandidateSelector
from priority_detector import detect_red_font_rows


@dataclass
class MatchResult:
    """매칭 결과"""
    no: int
    brand: str
    product_code: str
    product_name: str
    
    # GNC
    gnc_search_result: str = ""
    gnc_url: str = ""
    gnc_thumbnail: str = ""
    gnc_count: Optional[int] = None
    
    # 쿠팡 검색
    coupang_query: str = ""
    coupang_candidates_count: int = 0
    
    # 쿠팡 상품
    coupang_name: str = ""
    coupang_url: str = ""
    coupang_price: int = 0
    coupang_shipping: int = 0
    coupang_final_price: int = 0
    coupang_count: Optional[int] = None
    coupang_brand: str = ""
    coupang_rating: Optional[float] = None
    coupang_reviews: Optional[int] = None
    coupang_seller: str = ""
    
    # 후보 선택
    selection_confidence: str = ""
    selection_reason: str = ""
    
    # 이미지 비교
    image_match: str = ""
    image_confidence: str = ""
    image_reason: str = ""
    
    # 상태
    status: str = ""
    error_message: str = ""
    processed_at: str = ""


class ProductMatchingSystem:
    """GNC-쿠팡 자동 매칭"""
    
    def __init__(self, excel_path: str, gemini_api_key: Optional[str] = None, headless: bool = False):
        self.excel_path = excel_path
        self.headless = headless
        self.browser = None
        
        # AI 매처
        if gemini_api_key:
            self.image_matcher = ImageMatcher(gemini_api_key)
            self.candidate_selector = CandidateSelector(gemini_api_key)
            self.use_gemini = True
            print("✓ Gemini API 사용")
        else:
            self.image_matcher = None
            self.candidate_selector = None
            self.use_gemini = False
            print("⚠ Gemini API 없음")
        
        self.results: List[MatchResult] = []
        self.output_path = f"matching_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        # 기존 결과 파일 확인
        existing_files = sorted([f for f in os.listdir('.') if f.startswith('matching_results_') and f.endswith('.csv')])
        if existing_files:
            latest_file = existing_files[-1]
            print(f"⚠ 기존 결과 파일: {latest_file}")
            response = input("이어서 진행? (y/n): ").lower()
            if response == 'y':
                self.output_path = latest_file
                print(f"✓ 이어서 실행: {self.output_path}")
    
    def load_existing_results(self) -> set:
        """기존 처리된 NO 로드"""
        if os.path.exists(self.output_path):
            try:
                df = pd.read_csv(self.output_path, encoding='utf-8-sig')
                processed_nos = set(df['no'].tolist())
                print(f"✓ 기존 결과 {len(processed_nos)}개 로드")
                return processed_nos
            except:
                pass
        return set()
    
    def initialize_crawlers(self):
        """크롤러 초기화"""
        print("\n크롤러 초기화...")
        self.browser = CoupangBrowser(headless=self.headless)
        
        self.gnc_crawler = GNCCrawler(browser_manager=self.browser)
        self.coupang_crawler = CoupangCrawler(browser_manager=self.browser)
        print("✓ 준비 완료\n")
    
    def load_products(self, priority_numbers: Optional[List[int]] = None):
        """엑셀 로드"""
        print(f"엑셀 로드: {self.excel_path}")
        
        wb = load_workbook(self.excel_path)
        ws = wb.active
        
        priority_products = []
        normal_products = []
        
        headers = [cell.value for cell in ws[1]]
        
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data[header] = cell.value
            
            product_no = row_data.get('NO')
            
            if priority_numbers and product_no in priority_numbers:
                priority_products.append(row_data)
            else:
                normal_products.append(row_data)
        
        wb.close()
        
        print(f"✓ 우선순위: {len(priority_products)}개")
        print(f"✓ 일반: {len(normal_products)}개\n")
        
        return priority_products, normal_products
    
    def process_product(self, product_data: Dict) -> MatchResult:
        """개별 상품 처리"""
        no = product_data.get('NO', 0)
        brand = product_data.get('브랜드', '')
        product_code = product_data.get('상품코드', '')
        product_name = product_data.get('상품명', '')
        
        print(f"\n{'='*60}")
        print(f"[{no}] {brand} - {product_name}")
        print(f"{'='*60}")
        
        result = MatchResult(
            no=no,
            brand=brand,
            product_code=str(product_code),
            product_name=product_name,
            processed_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        
        try:
            # [1] GNC 검색
            print("\n[1] GNC 검색...")
            gnc = self.gnc_crawler.search_product(product_code)
            
            if not gnc:
                result.status = "실패"
                result.error_message = "GNC 검색 실패"
                return result
            
            result.gnc_search_result = gnc.product_name
            result.gnc_url = gnc.gnc_url
            result.gnc_thumbnail = gnc.thumbnail_url or ""
            result.gnc_count = gnc.count
            
            # [2] 쿠팡 쿼리
            print("\n[2] 쿼리 생성...")
            query = self._generate_query(gnc)
            result.coupang_query = query
            print(f"  {query}")
            
            # [3] 쿠팡 검색
            print("\n[3] 쿠팡 검색...")
            candidates = self.coupang_crawler.search_products(query, top_n=5)
            result.coupang_candidates_count = len(candidates)
            
            if not candidates:
                result.status = "실패"
                result.error_message = "쿠팡 검색 실패"
                return result
            
            # [4] 후보 선택 (엄격한 기준)
            print("\n[4] 후보 선택...")
            if self.candidate_selector:
                best, confidence, reason = self.candidate_selector.select_best_candidate(gnc, candidates)
                
                # 매칭 불가 처리
                if best is None or confidence == 'none':
                    result.status = "매칭불가"
                    result.error_message = "적합한 후보 없음"
                    result.selection_confidence = confidence
                    result.selection_reason = reason
                    print(f"  ✗ 매칭 불가")
                    return result
                
                print(f"  ✓ Gemini 선택 (신뢰도: {confidence})")
                result.selection_confidence = confidence
                result.selection_reason = reason
            else:
                # Gemini 없으면 매칭 불가
                result.status = "매칭불가"
                result.error_message = "Gemini API 없음"
                result.selection_confidence = "none"
                print(f"  ✗ Gemini API 없음 - 자동 매칭 불가")
                return result
            
            # [5] 상세 정보
            print("\n[5] 상세 정보...")
            detail = self.coupang_crawler.get_product_detail(best.url)

            # 기본 정보: 검색 결과 기준
            result.coupang_name = best.name
            result.coupang_url = best.url
            result.coupang_price = best.price
            result.coupang_shipping = best.shipping_fee
            result.coupang_final_price = best.final_price

            # 평점/리뷰: 우선 검색 결과 값을 넣어두고,
            # 상세 페이지에 있으면 그 값으로 덮어쓴다.
            result.coupang_rating = best.rating
            result.coupang_reviews = best.review_count

            if detail:
                # 판매자
                result.coupang_seller = detail.get('seller_name', '')

                # 상세페이지에 평점 정보가 있으면 우선 사용
                if detail.get('rating') is not None:
                    result.coupang_rating = detail['rating']

                # 상세페이지에 리뷰수 정보가 있으면 우선 사용
                if detail.get('review_count') is not None:
                    result.coupang_reviews = detail['review_count']

                # 썸네일 보정 (상세 이미지가 더 정확한 경우)
                if detail.get('thumbnail_url'):
                    best.thumbnail_url = detail['thumbnail_url']
            
            # [6] 이미지 비교
            print("\n[6] 이미지 비교...")
            print(f"  GNC 썸네일: {gnc.thumbnail_url or '-'}")
            print(f"  쿠팡 썸네일: {best.thumbnail_url or '-'}")
            if gnc.thumbnail_url and best.thumbnail_url and self.image_matcher:
                is_match, conf, reason = self.image_matcher.compare_images(
                    gnc.thumbnail_url, best.thumbnail_url
                )
                result.image_match = "일치" if is_match else "불일치"
                result.image_confidence = conf
                result.image_reason = reason
                
                # 종합 판정 (엄격한 기준)
                if confidence == 'high' and is_match and conf == 'high':
                    result.status = "성공"
                    print(f"  ✅ 높은 신뢰도 매칭")
                elif confidence == 'high' and is_match:
                    result.status = "성공"
                    print(f"  ✅ 매칭 성공")
                elif confidence == 'medium' and is_match and conf in ['high', 'medium']:
                    result.status = "검토필요"
                    result.error_message = "중간 신뢰도"
                    print(f"  ⚠️  검토 필요 (중간 신뢰도)")
                else:
                    result.status = "매칭불가"
                    result.error_message = "이미지 불일치 또는 낮은 신뢰도"
                    print(f"  ✗ 매칭 불가")
            else:
                # 이미지 비교 불가
                result.image_match = "비교불가"
                if confidence == 'high':
                    result.status = "검토필요"
                    result.error_message = "이미지 비교 불가"
                    print(f"  ⚠️  이미지 없음 - 검토 필요")
                else:
                    result.status = "매칭불가"
                    result.error_message = "이미지 없음 + 낮은 신뢰도"
                    print(f"  ✗ 이미지 없음 - 매칭 불가")
            
            print(f"\n✓ 완료: {result.status}")
            return result
            
        except Exception as e:
            print(f"\n✗ 오류: {e}")
            result.status = "오류"
            result.error_message = str(e)
            return result
    
    def _generate_query(self, gnc: GNCProduct) -> str:
        """쿠팡 검색 쿼리 생성 (GNC 검색 결과의 브랜드 + 상품명 그대로 사용)"""
        # GNC 검색결과에서 가져온 브랜드 + 상품명
        if gnc.brand:
            return f"{gnc.brand} {gnc.product_name}"
        return gnc.product_name
    
    def _select_best_fallback(self, gnc: GNCProduct, candidates: List[CoupangProduct]):
        """폴백 선택"""
        if gnc.count:
            matched = [c for c in candidates if c.count == gnc.count]
            if matched:
                return min(matched, key=lambda x: x.final_price)
        return min(candidates, key=lambda x: x.final_price)
    
    def save_results(self):
        """결과 저장"""
        if not self.results:
            return
        
        df = pd.DataFrame([asdict(r) for r in self.results])
        
        if os.path.exists(self.output_path):
            df.to_csv(self.output_path, mode='a', header=False, index=False, encoding='utf-8-sig')
        else:
            df.to_csv(self.output_path, index=False, encoding='utf-8-sig')
        
        self.results.clear()
        print(f"  💾 중간 저장 완료")
    
    def run(self, priority_numbers: Optional[List[int]] = None):
        """실행"""
        try:
            self.initialize_crawlers()
            priority, normal = self.load_products(priority_numbers)
            
            processed = self.load_existing_results()
            
            if processed:
                print(f"\n{'='*60}")
                print(f"이미 처리된 {len(processed)}개 건너뛰기")
                print(f"{'='*60}")
                
                if priority:
                    original = len(priority)
                    priority = [p for p in priority if p.get('NO') not in processed]
                    if original - len(priority) > 0:
                        print(f"✓ 우선순위: {original - len(priority)}개 건너뛰기")
                
                original = len(normal)
                normal = [p for p in normal if p.get('NO') not in processed]
                if original - len(normal) > 0:
                    print(f"✓ 일반: {original - len(normal)}개 건너뛰기")
            
            # 우선순위
            if priority:
                print(f"\n{'='*60}")
                print("우선순위 처리")
                print(f"{'='*60}")
                
                for idx, p in enumerate(priority, 1):
                    print(f"\n진행: {idx}/{len(priority)}")
                    result = self.process_product(p)
                    self.results.append(result)
                    
                    if idx % 10 == 0:
                        self.save_results()
                    time.sleep(2)
            
            # 일반
            print(f"\n{'='*60}")
            print("일반 처리")
            print(f"{'='*60}")
            
            for idx, p in enumerate(normal, 1):
                print(f"\n진행: {idx}/{len(normal)}")
                result = self.process_product(p)
                self.results.append(result)
                
                if idx % 10 == 0:
                    self.save_results()
                time.sleep(2)
            
            self.save_results()
            
            # 통계
            print(f"\n{'='*60}")
            print("완료")
            print(f"{'='*60}")
            
            df = pd.read_csv(self.output_path, encoding='utf-8-sig')
            success = len(df[df['status'] == '성공'])
            review = len(df[df['status'] == '검토필요'])
            no_match = len(df[df['status'] == '매칭불가'])
            failed = len(df[df['status'] == '실패'])
            error = len(df[df['status'] == '오류'])
            
            print(f"✅ 성공: {success}개")
            print(f"⚠️  검토필요: {review}개")
            print(f"❌ 매칭불가: {no_match}개")
            print(f"❌ 실패: {failed}개")
            print(f"❌ 오류: {error}개")
            
            total = len(df)
            print(f"\n📊 매칭률: {success / total * 100:.1f}%")
            print(f"📊 신뢰할 수 있는 결과: {(success + review) / total * 100:.1f}%")
        
        except KeyboardInterrupt:
            print("\n\n⚠️  사용자가 중단했습니다")
            print("지금까지 처리된 결과를 저장합니다...")
            self.save_results()
            print("✓ 저장 완료")
            
        finally:
            if self.browser:
                self.browser.close()


def main():
    excel_path = "GNC_상품_리스트_외산.xlsx"
    gemini_api_key = os.getenv('GEMINI_API_KEY')
    
    if not gemini_api_key:
        print("\n⚠️  경고: GEMINI_API_KEY가 설정되지 않았습니다")
        print("엄격한 매칭을 위해 Gemini API가 필수입니다")
        print("Gemini 없이 실행하면 모든 상품이 '매칭불가'로 처리됩니다")
        response = input("\n계속 진행하시겠습니까? (y/n): ").lower()
        if response != 'y':
            return
    
    # 우선순위 감지
    print("\n우선순위 감지 중...")
    priority_numbers = detect_red_font_rows(excel_path)
    
    if priority_numbers:
        print(f"✓ 빨간색 폰트 {len(priority_numbers)}개")
    else:
        priority_numbers = None
    
    print(f"\n{'='*60}")
    print("GNC-쿠팡 자동 매칭 (엄격한 기준)")
    print(f"{'='*60}")
    print("⚠️  매칭 정책:")
    print("  - 브랜드 불일치 → 매칭 불가")
    print("  - 성분 불일치 → 매칭 불가")
    print("  - 정수 차이 큼 → 매칭 불가")
    print("  - 의심스러우면 → 매칭 불가")
    print("  ✅ 잘못된 매칭보다 매칭 안 하는 게 낫습니다")
    print(f"{'='*60}\n")
    
    system = ProductMatchingSystem(
        excel_path=excel_path,
        gemini_api_key=gemini_api_key,
        headless=False
    )
    
    system.run(priority_numbers=priority_numbers)


if __name__ == '__main__':
    main()