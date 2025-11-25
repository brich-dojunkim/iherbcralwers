"""
빨간색 폰트 행 자동 감지
엑셀에서 폰트 색상이 빨간색(FFFF0000)인 행의 NO 번호 추출
"""

from openpyxl import load_workbook
from typing import List


def detect_red_font_rows(excel_path: str) -> List[int]:
    """
    빨간색 폰트 행의 NO 번호 추출
    
    Args:
        excel_path: 엑셀 파일 경로
        
    Returns:
        빨간색 폰트 행의 NO 번호 리스트
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # NO 컬럼 찾기
    headers = [cell.value for cell in ws[1]]
    no_col = headers.index('NO') + 1 if 'NO' in headers else 1
    
    red_rows = []
    
    # 모든 행 스캔
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        
        # 폰트 색상 확인
        if cell.font and cell.font.color:
            rgb = str(cell.font.color.rgb) if hasattr(cell.font.color, 'rgb') else ''
            
            # 빨간색 감지
            if 'FFFF0000' in rgb:
                no_cell = ws.cell(row=row_idx, column=no_col)
                if no_cell.value:
                    red_rows.append(int(no_cell.value))
    
    wb.close()
    return red_rows