#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
옵션 C: 하이브리드 구조 (측정값 vs 변화량 분리)
상위 헤더: 기본정보 | 측정값 | 변화량
하위 헤더: - | 순위(10/20,10/21,10/23), 가격(...), 할인율(...) | 순위변화(10/21,10/23), 가격변화(...), 할인율변화(...)

특징:
- 측정값과 변화량을 명확히 분리
- 미니멀 디자인
- 모든 변화량: 절대치
- 변화량 색상: 음수=파랑, 양수=빨강
"""

import sqlite3
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "/Users/brich/Desktop/iherb_price/coupang2/monitoring.db"
OUTPUT_DIR = "/Users/brich/Desktop/iherb_price/coupang2/outputs"

def generate_hybrid_format():
    """하이브리드 구조로 Wide Format 생성"""
    
    conn = sqlite3.connect(DB_PATH)
    
    print("\n" + "="*80)
    print("📊 옵션 C: 하이브리드 구조 (측정값 vs 변화량)")
    print("="*80)
    
    # 1. 데이터 로드
    query = """
    WITH daily_latest_snapshots AS (
        SELECT 
            c.id as category_id,
            c.name as category_name,
            DATE(ps.snapshot_time) as snapshot_date,
            MAX(ps.id) as latest_snapshot_id
        FROM page_snapshots ps
        JOIN categories c ON ps.category_id = c.id
        GROUP BY c.id, c.name, DATE(ps.snapshot_time)
    )
    SELECT 
        dls.category_name,
        dls.snapshot_date,
        prod.coupang_product_id,
        prod.product_name,
        prod.category_rank,
        prod.current_price,
        prod.discount_rate,
        prod.review_count,
        prod.rating_score,
        mr.iherb_upc,
        mr.iherb_part_number
    FROM daily_latest_snapshots dls
    JOIN product_states prod ON dls.latest_snapshot_id = prod.snapshot_id
    INNER JOIN matching_reference mr ON prod.coupang_product_id = mr.coupang_product_id
    WHERE mr.iherb_upc IS NOT NULL OR mr.iherb_part_number IS NOT NULL
    ORDER BY dls.category_name, prod.coupang_product_id, dls.snapshot_date
    """
    
    df_long = pd.read_sql_query(query, conn)
    conn.close()
    
    dates = sorted(df_long['snapshot_date'].unique())
    print(f"\n✅ 데이터 로드: {len(df_long):,}개 레코드, {len(dates)}일")
    
    # 2. 기본 정보
    base_info = df_long.groupby(['category_name', 'coupang_product_id']).agg({
        'product_name': 'first',
        'iherb_upc': 'first',
        'iherb_part_number': 'first',
        'review_count': 'last',
        'rating_score': 'last'
    }).reset_index()
    
    wide_df = base_info.copy()
    
    # 3. 날짜별 데이터 추가
    print(f"🔄 데이터 변환 중...")
    for date in dates:
        date_str = date[5:].replace('-', '/')
        date_data = df_long[df_long['snapshot_date'] == date][
            ['category_name', 'coupang_product_id', 'category_rank', 'current_price', 'discount_rate']
        ].copy()
        date_data.columns = ['category_name', 'coupang_product_id',
                            f'순위_{date_str}', f'가격_{date_str}', f'할인율_{date_str}']
        wide_df = wide_df.merge(date_data, on=['category_name', 'coupang_product_id'], how='left')
    
    # 4. 변화량 계산
    print(f"📊 변화량 계산 중...")
    for i in range(1, len(dates)):
        prev = dates[i-1][5:].replace('-', '/')
        curr = dates[i][5:].replace('-', '/')
        
        wide_df[f'순위변화_{curr}'] = (wide_df[f'순위_{prev}'] - wide_df[f'순위_{curr}']).round(0)
        wide_df[f'가격변화_{curr}'] = (wide_df[f'가격_{curr}'] - wide_df[f'가격_{prev}']).round(0)
        wide_df[f'할인율변화_{curr}'] = (wide_df[f'할인율_{curr}'] - wide_df[f'할인율_{prev}']).round(1)
    
    # 5. 컬럼 재정렬 (측정값 그룹 + 변화량 그룹)
    base_cols = ['category_name', 'coupang_product_id', 'product_name', 
                 'iherb_upc', 'iherb_part_number', 'review_count', 'rating_score']
    
    ordered_cols = base_cols.copy()
    
    # 측정값 그룹
    for d in dates:
        ordered_cols.append(f'순위_{d[5:].replace("-", "/")}')
    for d in dates:
        ordered_cols.append(f'가격_{d[5:].replace("-", "/")}')
    for d in dates:
        ordered_cols.append(f'할인율_{d[5:].replace("-", "/")}')
    
    # 변화량 그룹
    for i in range(1, len(dates)):
        ordered_cols.append(f'순위변화_{dates[i][5:].replace("-", "/")}')
    for i in range(1, len(dates)):
        ordered_cols.append(f'가격변화_{dates[i][5:].replace("-", "/")}')
    for i in range(1, len(dates)):
        ordered_cols.append(f'할인율변화_{dates[i][5:].replace("-", "/")}')
    
    wide_df = wide_df[ordered_cols]
    
    # 6. 멀티레벨 컬럼
    multi_columns = []
    base_names = {'category_name': '카테고리', 'coupang_product_id': '쿠팡상품ID',
                  'product_name': '상품명', 'iherb_upc': 'iHerb_UPC',
                  'iherb_part_number': 'iHerb_품번', 'review_count': '리뷰수',
                  'rating_score': '평점'}
    
    for col in base_cols:
        multi_columns.append(('기본정보', base_names[col]))
    
    # 측정값 그룹
    for d in dates:
        multi_columns.append(('측정값-순위', d[5:].replace('-', '/')))
    for d in dates:
        multi_columns.append(('측정값-가격', d[5:].replace('-', '/')))
    for d in dates:
        multi_columns.append(('측정값-할인율(%)', d[5:].replace('-', '/')))
    
    # 변화량 그룹
    for i in range(1, len(dates)):
        multi_columns.append(('변화량-순위', dates[i][5:].replace('-', '/')))
    for i in range(1, len(dates)):
        multi_columns.append(('변화량-가격', dates[i][5:].replace('-', '/')))
    for i in range(1, len(dates)):
        multi_columns.append(('변화량-할인율', dates[i][5:].replace('-', '/')))
    
    wide_df.columns = pd.MultiIndex.from_tuples(multi_columns)
    print(f"✅ 변환 완료: {len(wide_df):,}개 상품, {len(wide_df.columns)}개 컬럼")
    
    return dates, wide_df


def save_to_excel_minimal(dates, wide_df, output_path, option_name):
    """미니멀 디자인 Excel 저장"""
    
    print(f"\n💾 Excel 저장 중...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = '데이터'
    
    # 미니멀 스타일
    header_font = Font(bold=True, size=11)
    subheader_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    
    # 1행: 상위 헤더
    for col_idx, (level0, level1) in enumerate(wide_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=level0)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
    
    # 2행: 하위 헤더
    for col_idx, (level0, level1) in enumerate(wide_df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=level1)
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
    
    # 3행부터: 데이터 + 변화량 색상
    for row_idx, row_data in enumerate(wide_df.values, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 컬럼 정보
            col_name = wide_df.columns[col_idx - 1][0]  # 상위 헤더
            
            # 숫자 포맷 적용 (세자리수마다 콤마, 소수점 없음)
            if pd.notna(value) and isinstance(value, (int, float)):
                # 변화량 컬럼
                if '변화량' in col_name:
                    if value < 0:
                        cell.font = Font(color='0000FF')  # 음수: 파란색
                        cell.number_format = '#,##0'
                    elif value > 0:
                        cell.font = Font(color='FF0000')  # 양수: 빨간색
                        cell.number_format = '+#,##0'
                    else:
                        cell.number_format = '#,##0'
                else:
                    # 측정값 컬럼 (순위, 가격, 할인율)
                    cell.number_format = '#,##0'
    
    # 헤더 병합
    merge_start = 1
    prev_group = None
    for i, (level0, level1) in enumerate(wide_df.columns, start=1):
        if prev_group is not None and prev_group != level0:
            if merge_start < i - 1:
                ws.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=i - 1)
            merge_start = i
        prev_group = level0
    
    if merge_start < len(wide_df.columns):
        ws.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=len(wide_df.columns))
    
    # 그룹 간 구분선
    prev_group = None
    for col_idx, (level0, level1) in enumerate(wide_df.columns, start=1):
        if prev_group is not None and prev_group != level0:
            for row_idx in range(1, len(wide_df) + 3):
                cell = ws.cell(row=row_idx, column=col_idx - 1)
                cell.border = Border(right=Side(style='medium', color='CCCCCC'))
        prev_group = level0
    
    # 컬럼 너비 조정
    for col_idx in range(1, len(wide_df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'H3'
    
    wb.save(output_path)
    print(f"✅ 저장 완료: {output_path}")


def main():
    """메인 실행"""
    
    print("\n" + "="*80)
    print("🎯 옵션 C: 하이브리드 구조")
    print("="*80)
    print("상위 헤더: 기본정보 | 측정값 | 변화량")
    print("하위 헤더: - | 순위(10/20,10/21,...), 가격(...), 할인율(...) | 순위변화(10/21,...), 가격변화(...), 할인율변화(...)")
    print("\n특징:")
    print("  ✓ 측정값과 변화량을 명확히 분리")
    print("  ✓ 각 섹션 내에서 지표별 비교 용이")
    print("  ✓ 모든 변화량: 절대치")
    print("  ✓ 변화량 색상: 음수=파랑, 양수=빨강")
    print("  ✓ 미니멀 디자인")
    print("="*80)
    
    dates, wide_df = generate_hybrid_format()
    
    output_path = os.path.join(OUTPUT_DIR, "option_C_hybrid.xlsx")
    save_to_excel_minimal(dates, wide_df, output_path, "C")
    
    # CSV
    csv_path = os.path.join(OUTPUT_DIR, "option_C_hybrid.csv")
    flat_df = wide_df.copy()
    flat_df.columns = [f'{level0}_{level1}' if level0 != '기본정보' else level1 
                       for level0, level1 in flat_df.columns]
    flat_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"💾 CSV 저장: {csv_path}")
    
    print("\n" + "="*80)
    print("✅ 옵션 C 완료!")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()