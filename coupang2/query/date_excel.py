#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
최종 Price Agent Excel (2-Sheet)

Sheet 1: 전체 데이터 시계열 (프레젠테이션용) - date_excel 스타일
Sheet 2: 중복 UPC 상세 탐색 (그룹화) - 379개 UPC만
"""

import sqlite3
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from difflib import SequenceMatcher

DB_PATH = "/Users/brich/Desktop/iherb_price/coupang2/monitoring.db"
OUTPUT_DIR = "/Users/brich/Desktop/iherb_price/coupang2/outputs"


def calculate_name_similarity(name1, name2):
    """두 상품명의 유사도 계산 (0.0 ~ 1.0)"""
    if pd.isna(name1) or pd.isna(name2):
        return 0.0
    return SequenceMatcher(None, str(name1), str(name2)).ratio()


def select_representative_product(group_df):
    """
    UPC 그룹에서 대표 상품 선정
    
    우선순위:
    1. 소스 = 로켓직구
    2. 이름 유사도 최대 (평균)
    3. 리뷰수 내림차순
    4. 평점 내림차순
    5. 가격 오름차순
    """
    
    # 로켓직구 필터링
    rocket_products = group_df[group_df['source_type'] == 'rocket_direct'].copy()
    
    candidates = rocket_products if len(rocket_products) > 0 else group_df.copy()
    
    # 평균 유사도 계산
    for idx in candidates.index:
        product_name = candidates.loc[idx, 'product_name']
        similarities = []
        
        for other_idx in group_df.index:
            if idx != other_idx:
                other_name = group_df.loc[other_idx, 'product_name']
                sim = calculate_name_similarity(product_name, other_name)
                similarities.append(sim)
        
        candidates.loc[idx, 'avg_similarity'] = np.mean(similarities) if similarities else 0.0
    
    # 정렬
    candidates = candidates.sort_values(
        by=['avg_similarity', 'review_count', 'rating_score', 'current_price'],
        ascending=[False, False, False, True]
    )
    
    return candidates.index[0]


def generate_sheet1_all_products():
    """Sheet 1: 전체 데이터 시계열 (date_excel 스타일)"""
    
    conn = sqlite3.connect(DB_PATH)
    
    print("\n" + "="*80)
    print("📊 Sheet 1: 전체 데이터 시계열 생성 중...")
    print("="*80)
    
    # 데이터 로드
    query = """
    WITH daily_latest_snapshots AS (
        SELECT 
            c.name as category_name,
            s.source_type,
            DATE(snap.snapshot_time) as snapshot_date,
            MAX(snap.id) as latest_snapshot_id
        FROM snapshots snap
        JOIN sources s ON snap.source_id = s.id
        JOIN categories c ON snap.category_id = c.id
        GROUP BY c.name, s.source_type, DATE(snap.snapshot_time)
    )
    SELECT 
        dls.category_name,
        dls.source_type,
        dls.snapshot_date,
        mr.iherb_upc,
        mr.iherb_part_number,
        ps.vendor_item_id,
        ps.product_name,
        ps.category_rank,
        ps.current_price,
        ps.discount_rate,
        ps.review_count,
        ps.rating_score
    FROM daily_latest_snapshots dls
    JOIN product_states ps ON dls.latest_snapshot_id = ps.snapshot_id
    LEFT JOIN matching_reference mr ON ps.vendor_item_id = mr.vendor_item_id
    ORDER BY COALESCE(mr.iherb_upc, ps.vendor_item_id), 
             dls.category_name, dls.source_type, dls.snapshot_date
    """
    
    df_long = pd.read_sql_query(query, conn)
    conn.close()
    
    if df_long.empty:
        print("   ⚠️ 데이터 없음")
        return pd.DataFrame(), []
    
    dates = sorted(df_long['snapshot_date'].unique())
    print(f"   ✓ {len(df_long):,}개 레코드, {len(dates)}일")
    
    # 소스 라벨
    df_long['소스'] = df_long['source_type'].map({
        'rocket_direct': '로켓직구',
        'iherb_official': '아이허브'
    })
    
    # 기본 정보
    base = df_long.groupby(['category_name', 'source_type', 'iherb_upc', 'vendor_item_id']).agg({
        'product_name': 'first',
        'iherb_part_number': 'first',
        'review_count': 'last',
        'rating_score': 'last',
        '소스': 'first'
    }).reset_index()
    
    # 날짜별 데이터 추가
    for date in dates:
        date_str = date[5:].replace('-', '/')
        date_data = df_long[df_long['snapshot_date'] == date][
            ['category_name', 'source_type', 'iherb_upc', 'vendor_item_id',
             'category_rank', 'current_price', 'discount_rate', 'review_count']
        ].copy()
        
        date_data.columns = [
            'category_name', 'source_type', 'iherb_upc', 'vendor_item_id',
            f'{date_str}_순위', f'{date_str}_가격', f'{date_str}_할인율', f'{date_str}_리뷰수'
        ]
        
        base = base.merge(
            date_data,
            on=['category_name', 'source_type', 'iherb_upc', 'vendor_item_id'],
            how='left'
        )
    
    # 변화량 계산
    for i in range(1, len(dates)):
        prev = dates[i-1][5:].replace('-', '/')
        curr = dates[i][5:].replace('-', '/')
        
        # 순위 변화 = 이전 - 현재
        base[f'{curr}_순위변화'] = base[f'{prev}_순위'] - base[f'{curr}_순위']
        
        # 가격 변화 = 현재 - 이전
        base[f'{curr}_가격변화'] = base[f'{curr}_가격'] - base[f'{prev}_가격']
        
        # 할인율 변화 = 현재 - 이전
        base[f'{curr}_할인율변화'] = base[f'{curr}_할인율'] - base[f'{prev}_할인율']
    
    # 컬럼 정리
    base = base.rename(columns={
        'category_name': '카테고리',
        'iherb_upc': 'UPC',
        'iherb_part_number': '품번',
        'vendor_item_id': 'VendorItemID',
        'product_name': '상품명',
        'review_count': '최종리뷰수',
        'rating_score': '최종평점'
    })
    
    # 컬럼 순서 재정렬 (기본정보 + 날짜별 측정값 + 날짜별 변화량)
    base_cols = ['카테고리', '소스', 'UPC', '품번', 'VendorItemID', '상품명', '최종리뷰수', '최종평점']
    
    ordered_cols = base_cols.copy()
    
    # 날짜별 측정값
    for date in dates:
        d = date[5:].replace('-', '/')
        ordered_cols.extend([f'{d}_순위', f'{d}_가격', f'{d}_할인율', f'{d}_리뷰수'])
    
    # 날짜별 변화량
    for i in range(1, len(dates)):
        d = dates[i][5:].replace('-', '/')
        ordered_cols.extend([f'{d}_순위변화', f'{d}_가격변화', f'{d}_할인율변화'])
    
    # 존재하는 컬럼만 선택
    ordered_cols = [c for c in ordered_cols if c in base.columns]
    base = base[ordered_cols]
    
    print(f"   ✅ Wide Format 생성 완료: {len(base):,}개 행")
    
    return base, dates


def generate_sheet2_overlap_upc():
    """Sheet 2: 중복 UPC만 (379개 그룹화)"""
    
    conn = sqlite3.connect(DB_PATH)
    
    print("\n" + "="*80)
    print("📊 Sheet 2: 중복 UPC 상세 탐색 생성 중...")
    print("="*80)
    
    # 중복 UPC만 필터링
    query = """
    WITH daily_latest_snapshots AS (
        SELECT 
            c.name as category_name,
            s.source_type,
            DATE(snap.snapshot_time) as snapshot_date,
            MAX(snap.id) as latest_snapshot_id
        FROM snapshots snap
        JOIN sources s ON snap.source_id = s.id
        JOIN categories c ON snap.category_id = c.id
        GROUP BY c.name, s.source_type, DATE(snap.snapshot_time)
    ),
    overlap_upc AS (
        SELECT 
            mr.iherb_upc
        FROM product_states ps
        JOIN snapshots snap ON ps.snapshot_id = snap.id
        JOIN sources s ON snap.source_id = s.id
        INNER JOIN matching_reference mr ON ps.vendor_item_id = mr.vendor_item_id
        WHERE mr.iherb_upc IS NOT NULL
        GROUP BY mr.iherb_upc
        HAVING COUNT(DISTINCT s.source_type) = 2
    )
    SELECT 
        dls.category_name,
        dls.source_type,
        dls.snapshot_date,
        mr.iherb_upc,
        mr.iherb_part_number,
        ps.vendor_item_id,
        ps.product_name,
        ps.category_rank,
        ps.current_price,
        ps.discount_rate,
        ps.review_count,
        ps.rating_score
    FROM overlap_upc ou
    JOIN matching_reference mr ON ou.iherb_upc = mr.iherb_upc
    JOIN product_states ps ON mr.vendor_item_id = ps.vendor_item_id
    JOIN daily_latest_snapshots dls ON ps.snapshot_id = dls.latest_snapshot_id
    ORDER BY mr.iherb_upc, dls.category_name, dls.source_type, dls.snapshot_date
    """
    
    df_long = pd.read_sql_query(query, conn)
    conn.close()
    
    if df_long.empty:
        print("   ⚠️ 중복 UPC 없음")
        return pd.DataFrame(), []
    
    dates = sorted(df_long['snapshot_date'].unique())
    print(f"   ✓ {len(df_long):,}개 레코드, {len(dates)}일")
    print(f"   ✓ 중복 UPC: {df_long['iherb_upc'].nunique():,}개")
    
    # 소스 라벨
    df_long['소스'] = df_long['source_type'].map({
        'rocket_direct': '로켓직구',
        'iherb_official': '아이허브'
    })
    
    # 최신 스냅샷으로 대표 선정
    latest_date = dates[-1]
    df_latest = df_long[df_long['snapshot_date'] == latest_date].copy()
    
    print(f"   🔍 대표 상품 선정 중... (기준: {latest_date})")
    
    # UPC별 대표 선정
    representative_mapping = {}
    
    for upc in df_latest['iherb_upc'].unique():
        if pd.notna(upc):
            group_df = df_latest[df_latest['iherb_upc'] == upc].copy()
            
            if len(group_df) > 0:
                rep_idx = select_representative_product(group_df)
                representative_mapping[upc] = group_df.loc[rep_idx, 'vendor_item_id']
    
    print(f"   ✓ {len(representative_mapping):,}개 그룹의 대표 선정 완료")
    
    # Wide Format 생성
    base = df_long.groupby(['category_name', 'source_type', 'iherb_upc', 'vendor_item_id']).agg({
        'product_name': 'first',
        'iherb_part_number': 'first',
        'review_count': 'last',
        'rating_score': 'last',
        '소스': 'first'
    }).reset_index()
    
    # 날짜별 데이터
    for date in dates:
        date_str = date[5:].replace('-', '/')
        date_data = df_long[df_long['snapshot_date'] == date][
            ['category_name', 'source_type', 'iherb_upc', 'vendor_item_id',
             'category_rank', 'current_price', 'discount_rate', 'review_count']
        ].copy()
        
        date_data.columns = [
            'category_name', 'source_type', 'iherb_upc', 'vendor_item_id',
            f'{date_str}_순위', f'{date_str}_가격', f'{date_str}_할인율', f'{date_str}_리뷰수'
        ]
        
        base = base.merge(
            date_data,
            on=['category_name', 'source_type', 'iherb_upc', 'vendor_item_id'],
            how='left'
        )
    
    # 변화량 계산
    for i in range(1, len(dates)):
        prev = dates[i-1][5:].replace('-', '/')
        curr = dates[i][5:].replace('-', '/')
        
        base[f'{curr}_순위변화'] = base[f'{prev}_순위'] - base[f'{curr}_순위']
        base[f'{curr}_가격변화'] = base[f'{curr}_가격'] - base[f'{prev}_가격']
        base[f'{curr}_할인율변화'] = base[f'{curr}_할인율'] - base[f'{prev}_할인율']
    
    # 그룹화 준비
    base['_upc_sort'] = base['iherb_upc']
    base['_group_order'] = 999
    base['_is_representative'] = False
    
    # 대표 상품 표시
    for upc, rep_vendor_id in representative_mapping.items():
        mask = (base['iherb_upc'] == upc) & (base['vendor_item_id'] == rep_vendor_id)
        base.loc[mask, '_group_order'] = 0
        base.loc[mask, '_is_representative'] = True
    
    # 나머지 상품 순서
    for upc in representative_mapping.keys():
        group_mask = base['iherb_upc'] == upc
        other_indices = base[group_mask & (base['_group_order'] != 0)].index
        
        if len(other_indices) > 0:
            group_subset = base.loc[other_indices].sort_values(
                ['source_type', 'product_name'],
                ascending=[False, True]
            )
            
            for i, idx in enumerate(group_subset.index, start=1):
                base.loc[idx, '_group_order'] = i
    
    # 정렬
    base = base.sort_values(['category_name', '_upc_sort', '_group_order'])
    
    # 컬럼 정리
    base = base.rename(columns={
        'category_name': '카테고리',
        'iherb_upc': 'UPC',
        'iherb_part_number': '품번',
        'vendor_item_id': 'VendorItemID',
        'product_name': '상품명',
        'review_count': '최종리뷰수',
        'rating_score': '최종평점'
    })
    
    # 컬럼 순서
    hidden_cols = ['_upc_sort', '_group_order', '_is_representative']
    base_cols = ['카테고리', '소스', 'UPC', '품번', 'VendorItemID', '상품명', '최종리뷰수', '최종평점']
    
    ordered_cols = hidden_cols + base_cols.copy()
    
    # 날짜별 측정값 + 변화량
    for date in dates:
        d = date[5:].replace('-', '/')
        ordered_cols.extend([f'{d}_순위', f'{d}_가격', f'{d}_할인율', f'{d}_리뷰수'])
    
    for i in range(1, len(dates)):
        d = dates[i][5:].replace('-', '/')
        ordered_cols.extend([f'{d}_순위변화', f'{d}_가격변화', f'{d}_할인율변화'])
    
    ordered_cols = [c for c in ordered_cols if c in base.columns]
    base = base[ordered_cols]
    
    print(f"   ✅ 그룹화 데이터 생성 완료: {len(base):,}개 행")
    
    return base, dates


def save_to_excel_date_style(df_sheet1, dates1, df_sheet2, dates2, output_path):
    """date_excel 스타일로 Excel 저장"""
    
    print("\n" + "="*80)
    print("💾 Excel 파일 저장 중 (date_excel 스타일)...")
    print("="*80)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # 스타일
    header_font = Font(bold=True, size=11)
    subheader_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    rep_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
    rep_font = Font(bold=True, size=10)
    
    # ===== Sheet 1: 전체 데이터 =====
    print("\n📄 Sheet 1: 전체 데이터...")
    ws1 = wb.create_sheet("전체데이터_시계열")
    
    # 멀티레벨 헤더 구성
    base_cols_count = 8  # 기본정보 컬럼 수
    
    # 1행: 상위 헤더
    ws1.cell(1, 1, "기본정보")
    
    col_idx = base_cols_count + 1
    for date in dates1:
        date_str = date[5:].replace('-', '/')
        ws1.cell(1, col_idx, date_str)
        col_idx += 4  # 순위, 가격, 할인율, 리뷰수
    
    for i in range(1, len(dates1)):
        date_str = dates1[i][5:].replace('-', '/')
        ws1.cell(1, col_idx, f"{date_str}_변화")
        col_idx += 3  # 순위변화, 가격변화, 할인율변화
    
    # 2행: 하위 헤더
    for c_idx, col_name in enumerate(df_sheet1.columns, 1):
        if col_name in ['카테고리', '소스', 'UPC', '품번', 'VendorItemID', '상품명', '최종리뷰수', '최종평점']:
            ws1.cell(2, c_idx, col_name)
        elif '_' in col_name:
            parts = col_name.rsplit('_', 1)
            ws1.cell(2, c_idx, parts[1])
    
    # 헤더 스타일
    for row in [1, 2]:
        for col in range(1, len(df_sheet1.columns) + 1):
            cell = ws1.cell(row, col)
            cell.font = header_font if row == 1 else subheader_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 상위 헤더 병합
    merge_start = 1
    prev_header = ws1.cell(1, 1).value
    
    for col in range(2, len(df_sheet1.columns) + 2):
        current_header = ws1.cell(1, col).value if col <= len(df_sheet1.columns) else None
        
        if current_header != prev_header or col == len(df_sheet1.columns) + 1:
            if merge_start < col - 1:
                ws1.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=col - 1)
            merge_start = col
            prev_header = current_header
    
    # 3행부터: 데이터
    for r_idx, row_data in enumerate(df_sheet1.values, start=3):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            
            col_name = df_sheet1.columns[c_idx - 1]
            
            # 변화량 색상
            if '변화' in col_name and pd.notna(value) and isinstance(value, (int, float)):
                if value < 0:
                    cell.font = Font(color='0000FF')
                elif value > 0:
                    cell.font = Font(color='FF0000')
    
    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = 'I3'
    
    print(f"   ✓ {len(df_sheet1):,}개 행 작성")
    
    # ===== Sheet 2: 중복 UPC 그룹화 =====
    print("\n📄 Sheet 2: 중복 UPC 그룹화...")
    ws2 = wb.create_sheet("중복UPC_그룹화")
    
    # 멀티레벨 헤더 (Sheet 1과 동일)
    ws2.cell(1, 4, "기본정보")  # 숨김 컬럼 3개 제외
    
    col_idx = 11  # 숨김 3개 + 기본정보 8개
    for date in dates2:
        date_str = date[5:].replace('-', '/')
        ws2.cell(1, col_idx, date_str)
        col_idx += 4
    
    for i in range(1, len(dates2)):
        date_str = dates2[i][5:].replace('-', '/')
        ws2.cell(1, col_idx, f"{date_str}_변화")
        col_idx += 3
    
    # 2행: 하위 헤더
    for c_idx, col_name in enumerate(df_sheet2.columns, 1):
        if col_name.startswith('_'):
            ws2.cell(2, c_idx, col_name)
        elif col_name in ['카테고리', '소스', 'UPC', '품번', 'VendorItemID', '상품명', '최종리뷰수', '최종평점']:
            ws2.cell(2, c_idx, col_name)
        elif '_' in col_name:
            parts = col_name.rsplit('_', 1)
            ws2.cell(2, c_idx, parts[1])
    
    # 헤더 스타일
    for row in [1, 2]:
        for col in range(4, len(df_sheet2.columns) + 1):  # 숨김 제외
            cell = ws2.cell(row, col)
            cell.font = header_font if row == 1 else subheader_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 상위 헤더 병합
    merge_start = 4
    prev_header = ws2.cell(1, 4).value
    
    for col in range(5, len(df_sheet2.columns) + 2):
        current_header = ws2.cell(1, col).value if col <= len(df_sheet2.columns) else None
        
        if current_header != prev_header or col == len(df_sheet2.columns) + 1:
            if merge_start < col - 1:
                ws2.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=col - 1)
            merge_start = col
            prev_header = current_header
    
    # 데이터 & 그룹화
    for r_idx, row_data in enumerate(df_sheet2.values, start=3):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            
            is_rep = df_sheet2.iloc[r_idx - 3]['_is_representative']
            col_name = df_sheet2.columns[c_idx - 1]
            
            # 대표 상품 스타일
            if is_rep and c_idx > 3:
                cell.fill = rep_fill
                cell.font = rep_font
            
            # 변화량 색상
            if '변화' in col_name and pd.notna(value) and isinstance(value, (int, float)):
                if value < 0:
                    cell.font = Font(color='0000FF', bold=is_rep)
                elif value > 0:
                    cell.font = Font(color='FF0000', bold=is_rep)
    
    # 숨김 컬럼
    ws2.column_dimensions['A'].hidden = True
    ws2.column_dimensions['B'].hidden = True
    ws2.column_dimensions['C'].hidden = True
    
    # 그룹화
    print("   🔧 그룹화 적용 중...")
    prev_upc = None
    group_start = 3
    group_count = 0
    
    for idx in range(len(df_sheet2)):
        excel_row = idx + 3
        current_upc = df_sheet2.iloc[idx]['_upc_sort']
        
        if prev_upc != current_upc:
            if prev_upc is not None and group_start < excel_row - 1:
                ws2.row_dimensions.group(
                    group_start + 1,
                    excel_row - 1,
                    outline_level=1,
                    hidden=False
                )
                group_count += 1
            
            group_start = excel_row
            prev_upc = current_upc
    
    # 마지막 그룹
    if group_start < len(df_sheet2) + 2:
        ws2.row_dimensions.group(
            group_start + 1,
            len(df_sheet2) + 2,
            outline_level=1,
            hidden=False
        )
        group_count += 1
    
    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = 'K3'
    
    print(f"   ✓ {len(df_sheet2):,}개 행, {group_count:,}개 그룹")
    
    # 저장
    wb.save(output_path)
    print(f"\n✅ 저장 완료: {output_path}")


def main():
    """메인 실행"""
    
    print("\n" + "="*80)
    print("🎯 최종 Price Agent Excel (2-Sheet)")
    print("="*80)
    print("\n구조:")
    print("  📊 Sheet 1: 전체 데이터 시계열 (프레젠테이션)")
    print("  📊 Sheet 2: 중복 UPC 그룹화 (379개)")
    print("\n스타일:")
    print("  ✓ date_excel 형태 (상위 헤더)")
    print("  ✓ 변화량 색상 (음수=파랑, 양수=빨강)")
    print("  ✓ 데이터 있는 그대로 (Round 없음)")
    print("="*80)
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Sheet 생성
    df_sheet1, dates1 = generate_sheet1_all_products()
    df_sheet2, dates2 = generate_sheet2_overlap_upc()
    
    if df_sheet1.empty or df_sheet2.empty:
        print("\n❌ 데이터 생성 실패")
        return
    
    # Excel 저장
    output_path = os.path.join(OUTPUT_DIR, "price_agent_final.xlsx")
    save_to_excel_date_style(df_sheet1, dates1, df_sheet2, dates2, output_path)
    
    # 통계
    print("\n" + "="*80)
    print("📊 최종 통계")
    print("="*80)
    print(f"\n✅ Sheet 1 (전체 데이터)")
    print(f"   - 총 행: {len(df_sheet1):,}개")
    print(f"   - 날짜: {len(dates1)}일")
    
    print(f"\n✅ Sheet 2 (중복 UPC)")
    print(f"   - 총 행: {len(df_sheet2):,}개")
    print(f"   - 중복 UPC: {df_sheet2['UPC'].nunique():,}개")
    print(f"   - 대표 상품: {df_sheet2['_is_representative'].sum():,}개")
    
    print("\n" + "="*80)
    print("✅ 완료!")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()