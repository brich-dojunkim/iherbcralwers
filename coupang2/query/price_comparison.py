#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
로켓직구 전체 상품 기준 가격 비교 테이블 생성
- 로켓직구 전체 상품 표시 (카테고리별 순위 기준 정렬)
- 매칭되는 아이허브 상품이 있으면 해당 정보 표시
- 매칭 조건: UPC 일치 + 상품명 마지막 개수 일치
- 일자별 시트 구성
- 링크는 "Link" 텍스트에 하이퍼링크로 표현
"""

import sqlite3
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime
import os
import re

# DB 경로 설정 (실제 경로로 수정 필요)
DB_PATH = "/Users/brich/Desktop/iherb_price/coupang2/monitoring.db"
OUTPUT_DIR = "./output"


def extract_quantity_from_product_name(product_name):
    """
    상품명 마지막에서 'X개' 패턴 추출
    
    예시:
    - "비타민 D3, 125mcg, 2개" → 2
    - "오메가3, 1개" → 1
    - "프로틴 파우더" → None (개수 없음)
    
    Args:
        product_name: 상품명
    
    Returns:
        int or None: 추출된 개수
    """
    if not product_name:
        return None
    
    # 마지막 부분에서 "숫자개" 패턴 찾기
    # 예: "2개", "10개", "120개"
    match = re.search(r'(\d+)개\s*$', product_name.strip())
    
    if match:
        return int(match.group(1))
    
    return None


def get_available_dates():
    """사용 가능한 날짜 목록 조회"""
    conn = sqlite3.connect(DB_PATH)
    query = """
    SELECT DISTINCT DATE(snapshot_time) as date
    FROM snapshots
    ORDER BY date DESC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df['date'].tolist()


def extract_price_comparison_data(target_date=None):
    """
    로켓직구 전체 상품 기준 가격 비교 데이터 추출
    매칭되는 아이허브 상품이 있으면 해당 정보 추가
    
    Args:
        target_date: 대상 날짜 (None이면 최신 날짜)
    
    Returns:
        DataFrame: 가격 비교 테이블
    """
    
    conn = sqlite3.connect(DB_PATH)
    
    # 날짜 결정
    if target_date is None:
        target_date = conn.execute("""
            SELECT DATE(MAX(snapshot_time)) FROM snapshots
        """).fetchone()[0]
    
    print(f"\n📅 처리 날짜: {target_date}")
    
    # ─────────────────────────────────────────────────────────
    # A안 적용: 조인 단계에서의 중복 차단 (SQL CTE 수정)
    #  - latest_snapshots: 같은 날짜 내 source_id, category_id 별 최신 스냅샷 1개만 선택
    #  - mr_dedup: matching_reference 중복 제거
    #  - iherb_products: UPC+카테고리 단위 대표 1행(최상위 랭크)만 남김
    # ─────────────────────────────────────────────────────────
    query = """
    WITH latest_snapshots AS (
    SELECT s2.id AS snapshot_id,
            src.source_type,
            src.display_name AS source_name,
            cat.name AS category
    FROM (
        SELECT MAX(s1.id) AS snapshot_id
        FROM snapshots s1
        WHERE DATE(s1.snapshot_time) = ?
        GROUP BY s1.source_id, s1.category_id
    ) pick
    JOIN snapshots  s2  ON s2.id = pick.snapshot_id
    JOIN sources    src ON src.id = s2.source_id
    JOIN categories cat ON cat.id = s2.category_id
    ),
    mr_dedup AS (
    SELECT vendor_item_id,
            MAX(iherb_upc)         AS iherb_upc,
            MAX(iherb_part_number) AS iherb_part_number
    FROM matching_reference
    GROUP BY vendor_item_id
    ),
    rocket_products AS (
    SELECT 
        ps.vendor_item_id,
        ps.product_name,
        ps.current_price,
        ps.original_price,
        ps.discount_rate,
        ps.category_rank,
        ps.rating_score,
        ps.review_count,
        ps.product_url,
        mr.iherb_upc         AS upc,
        mr.iherb_part_number AS iherb_part_number,
        ls.category          AS rocket_category
    FROM product_states ps
    JOIN latest_snapshots ls ON ps.snapshot_id = ls.snapshot_id
    LEFT JOIN mr_dedup mr     ON ps.vendor_item_id = mr.vendor_item_id
    WHERE ls.source_type = 'rocket_direct'
    ),
    iherb_products_raw AS (
    SELECT 
        ps.vendor_item_id,
        ps.product_name,
        ps.current_price,
        ps.original_price,
        ps.discount_rate,
        ps.category_rank,
        ps.rating_score,
        ps.review_count,
        ps.product_url,
        mr.iherb_upc         AS upc,
        mr.iherb_part_number AS iherb_part_number,
        ls.category          AS iherb_category
    FROM product_states ps
    JOIN latest_snapshots ls ON ps.snapshot_id = ls.snapshot_id
    LEFT JOIN mr_dedup mr     ON ps.vendor_item_id = mr.vendor_item_id
    WHERE ls.source_type = 'iherb_official'
    ),
    -- UPC 기준 대표 1행(가장 높은 노출: 낮은 카테고리 랭크 → 낮은 가격)
    iherb_products AS (
    SELECT *
    FROM (
        SELECT
        *,
        ROW_NUMBER() OVER (
            PARTITION BY upc
            ORDER BY category_rank ASC NULLS LAST, current_price ASC NULLS LAST
        ) AS rn
        FROM iherb_products_raw
    )
    WHERE rn = 1
    )
    SELECT 
        r.rocket_category       AS category,
        r.category_rank         AS rocket_rank,
        r.upc,
        r.iherb_part_number,

        -- 로켓직구
        r.vendor_item_id        AS rocket_vendor_id,
        r.product_name          AS rocket_product_name,
        r.current_price         AS rocket_price,
        r.original_price        AS rocket_original_price,
        r.discount_rate         AS rocket_discount_rate,
        r.rating_score          AS rocket_rating,
        r.review_count          AS rocket_reviews,
        r.product_url           AS rocket_url,

        -- 아이허브 공식 (UPC만 일치하면 매칭; 카테고리는 아이허브 쪽 실제 카테고리)
        i.vendor_item_id        AS iherb_vendor_id,
        i.product_name          AS iherb_product_name,
        i.current_price         AS iherb_price,
        i.original_price        AS iherb_original_price,
        i.discount_rate         AS iherb_discount_rate,
        i.category_rank         AS iherb_rank,
        i.rating_score          AS iherb_rating,
        i.review_count          AS iherb_reviews,
        i.product_url           AS iherb_url,
        i.iherb_category        AS iherb_category

    FROM rocket_products r
    LEFT JOIN iherb_products i 
    ON r.upc = i.upc
    ORDER BY r.rocket_category, r.category_rank
    """
    
    df = pd.read_sql_query(query, conn, params=(target_date,))
    conn.close()
    
    if df.empty:
        print("⚠️ 데이터가 없습니다.")
        return df

    # ─────────────────────────────────────────────────────────
    # 숫자형 컬럼 강제 변환(텍스트/쉼표/NULL 방지)
    # ─────────────────────────────────────────────────────────
    num_cols = [
        'rocket_price','rocket_original_price','rocket_discount_rate',
        'rocket_rating','rocket_reviews','rocket_rank',
        'iherb_price','iherb_original_price','iherb_discount_rate',
        'iherb_rating','iherb_reviews','iherb_rank'
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce')
    
    # 상품명에서 개수 추출
    df['rocket_quantity'] = df['rocket_product_name'].apply(extract_quantity_from_product_name)
    df['iherb_quantity'] = df['iherb_product_name'].apply(extract_quantity_from_product_name)
    
    # 개수가 다르면 매칭 해제
    def should_match(row):
        # UPC가 없으면 매칭 안 됨
        if pd.isna(row['upc']):
            return False
        
        # 아이허브 상품이 없으면 매칭 안 됨
        if pd.isna(row['iherb_vendor_id']):
            return False
        
        # 둘 다 개수가 있는 경우: 개수가 같아야 매칭
        if pd.notna(row['rocket_quantity']) and pd.notna(row['iherb_quantity']):
            return row['rocket_quantity'] == row['iherb_quantity']
        
        # 둘 중 하나만 개수가 있는 경우: 매칭 안 됨
        if pd.notna(row['rocket_quantity']) or pd.notna(row['iherb_quantity']):
            return False
        
        # 둘 다 개수가 없는 경우: UPC만으로 매칭
        return True
    
    # 매칭 여부 판단
    df['is_matched'] = df.apply(should_match, axis=1)
    
    # 매칭 안 되는 행은 아이허브 정보 제거
    iherb_columns = ['iherb_vendor_id', 'iherb_product_name', 'iherb_price', 
                     'iherb_original_price', 'iherb_discount_rate', 'iherb_rank',
                     'iherb_rating', 'iherb_reviews', 'iherb_url', 'iherb_category']
    for col in iherb_columns:
        df.loc[~df['is_matched'], col] = None
    
    # 비교 지표 계산 (매칭된 경우만)
    df['price_diff'] = None
    df['price_diff_pct'] = None
    df['cheaper_source'] = None
    df['discount_diff'] = None
    df['rating_diff'] = None
    df['review_diff'] = None
    
    matched_mask = df['is_matched']
    rp = df.loc[matched_mask, 'rocket_price']
    ip = df.loc[matched_mask, 'iherb_price']
    
    df.loc[matched_mask, 'price_diff'] = ip - rp
    df.loc[matched_mask, 'price_diff_pct'] = ((ip - rp) / rp.replace(0, np.nan) * 100).round(1)
    df.loc[matched_mask, 'cheaper_source'] = np.where(
        df.loc[matched_mask, 'price_diff'] > 0, '로켓직구',
        np.where(df.loc[matched_mask, 'price_diff'] < 0, '아이허브', '동일')
    )
    df.loc[matched_mask, 'discount_diff'] = df.loc[matched_mask, 'iherb_discount_rate'] - df.loc[matched_mask, 'rocket_discount_rate']
    df.loc[matched_mask, 'rating_diff']   = (df.loc[matched_mask, 'iherb_rating'] - df.loc[matched_mask, 'rocket_rating']).round(1)
    df.loc[matched_mask, 'review_diff']   = df.loc[matched_mask, 'iherb_reviews'] - df.loc[matched_mask, 'rocket_reviews']
    
    # 임시 컬럼 제거
    df = df.drop(['rocket_quantity', 'iherb_quantity', 'is_matched'], axis=1)
    
    print(f"✅ 총 {len(df)}개 로켓직구 상품")
    print(f"   - 매칭된 아이허브 상품: {df['iherb_vendor_id'].notna().sum()}개")
    print(f"   - 카테고리별:")
    for cat, count in df['category'].value_counts().items():
        matched = df[(df['category'] == cat) & (df['iherb_vendor_id'].notna())].shape[0]
        print(f"      • {cat}: 전체 {count}개 / 매칭 {matched}개")
    
    return df


def create_excel_report(date_data_dict, output_path):
    """
    엑셀 리포트 생성 (일자별 시트)
    - 상위 헤더(1행): [기본 정보(1~4)] [로켓직구(5~12)] [아이허브 공식(13~22)] [분석(23~28)]
      * 아이허브 공식 그룹에 '아이허브_카테고리' 포함
      * 분석 그룹에 '리뷰수차이' 포함
    - 하위 헤더(2행): 한 줄(shrink-to-fit), 데이터(3행~): 줄바꿈 없음
    - 그룹 경계(4, 12, 22열)에 흰색 두꺼운 세로 구분선
    - 오토필터 적용 후, 열 너비 자동 산정(오토핏 근사)
    """
    if not date_data_dict:
        print("❌ 데이터가 없어 엑셀 생성을 건너뜁니다.")
        return

    # 1) 컬럼 정의 (총 28개)
    columns_order = [
        # 기본 정보 (1~4)
        ('category', '카테고리'),
        ('rocket_rank', '로켓_순위'),
        ('upc', 'UPC'),
        ('iherb_part_number', '품번'),

        # 로켓직구 (5~12)
        ('rocket_vendor_id', '로켓_상품ID'),
        ('rocket_product_name', '로켓_제품명'),
        ('rocket_price', '로켓_가격'),
        ('rocket_original_price', '로켓_원가'),
        ('rocket_discount_rate', '로켓_할인율(%)'),
        ('rocket_rating', '로켓_평점'),
        ('rocket_reviews', '로켓_리뷰수'),
        ('rocket_url', '로켓_링크'),

        # 아이허브 공식 (13~22)  ← iherb_category 포함
        ('iherb_vendor_id', '아이허브_상품ID'),
        ('iherb_category', '아이허브_카테고리'),
        ('iherb_product_name', '아이허브_제품명'),
        ('iherb_price', '아이허브_가격'),
        ('iherb_original_price', '아이허브_원가'),
        ('iherb_discount_rate', '아이허브_할인율(%)'),
        ('iherb_rank', '아이허브_순위'),
        ('iherb_rating', '아이허브_평점'),
        ('iherb_reviews', '아이허브_리뷰수'),
        ('iherb_url', '아이허브_링크'),

        # 분석 (23~28)
        ('price_diff', '가격차이(원)'),
        ('price_diff_pct', '가격차이율(%)'),
        ('cheaper_source', '더_저렴한_곳'),
        ('discount_diff', '할인율차이(%p)'),
        ('rating_diff', '평점차이'),
        ('review_diff', '리뷰수차이'),
    ]

    # 2) 상위 헤더(병합 범위) & 그룹 경계
    super_headers = [
        ("기본 정보", 1, 4),
        ("로켓직구", 5, 12),
        ("아이허브 공식", 13, 22),  # 아이허브_카테고리 포함
        ("분석", 23, 28),            # 리뷰수차이 포함
    ]
    group_boundaries = [4, 12, 22]   # 각 그룹의 마지막 열

    # 3) 먼저 시트에 데이터 기록
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for date_str in sorted(date_data_dict.keys(), reverse=True):
            df = date_data_dict[date_str]
            if df.empty:
                continue
            df_export = df[[c[0] for c in columns_order]].copy()
            df_export.columns = [c[1] for c in columns_order]
            df_export.to_excel(writer, sheet_name=date_str, index=False, startrow=0)

    # 4) 스타일링
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_path)

    # 하위 헤더(2행)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=True)

    # 상위 헤더(1행)
    super_fill = PatternFill(start_color="244062", end_color="244062", fill_type="solid")
    super_font = Font(bold=True, color="FFFFFF", size=12)
    super_alignment = Alignment(horizontal='center', vertical='center')

    # 기본 얇은 테두리
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # 그룹 경계용 흰색 두꺼운 선
    thick_white_side = Side(style='thick', color='FFFFFF')

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # (a) 상위 헤더를 위한 1행 삽입
        ws.insert_rows(1)

        max_col = ws.max_column

        # (b) 하위 헤더(2행) 스타일
        for cell in ws[2]:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace("\n", " ").strip()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # (c) 상위 헤더 병합 및 스타일
        for title, c_start, c_end in super_headers:
            ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
            cell = ws.cell(row=1, column=c_start)
            cell.value = title
            cell.fill = super_fill
            cell.font = super_font
            cell.alignment = super_alignment

        # (d) 데이터 영역: 줄바꿈 없음 + 테두리
        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=False)

        # (e) 조건부 서식 & 링크
        header_row_values = [cell.value for cell in ws[2]]
        def col_idx_of(name): return header_row_values.index(name) + 1

        try:
            price_diff_col  = col_idx_of('가격차이(원)')
            cheaper_src_col = col_idx_of('더_저렴한_곳')
            rocket_url_col  = col_idx_of('로켓_링크')
            iherb_url_col   = col_idx_of('아이허브_링크')
        except ValueError:
            pass
        else:
            for row_idx in range(3, ws.max_row + 1):
                cheaper_value = ws.cell(row=row_idx, column=cheaper_src_col).value
                if cheaper_value == '아이허브':
                    ws.cell(row=row_idx, column=price_diff_col).fill = green_fill
                elif cheaper_value == '로켓직구':
                    ws.cell(row=row_idx, column=price_diff_col).fill = red_fill

                # 로켓 링크
                rc = ws.cell(row=row_idx, column=rocket_url_col)
                url_r = rc.value
                if url_r and str(url_r).strip():
                    rc.value = "Link"
                    rc.hyperlink = str(url_r)
                    rc.font = Font(color="0563C1", underline="single")
                    rc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

                # 아이허브 링크
                ic = ws.cell(row=row_idx, column=iherb_url_col)
                url_i = ic.value
                if url_i and str(url_i).strip():
                    ic.value = "Link"
                    ic.hyperlink = str(url_i)
                    ic.font = Font(color="0563C1", underline="single")
                    ic.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

        # (f) 그룹 경계(흰색 두꺼운 세로줄) 적용: 1행~마지막 행
        last_row = ws.max_row
        for boundary_col in group_boundaries:
            for r in range(1, last_row + 1):
                c = ws.cell(row=r, column=boundary_col)
                c.border = Border(
                    left=c.border.left,
                    right=thick_white_side,   # 오른쪽 경계선만 두껍게
                    top=c.border.top,
                    bottom=c.border.bottom
                )

        # (g) 틀 고정 & 필터 (먼저 적용)
        ws.freeze_panes = 'E3'  # 상위(1) + 하위(2) 아래 + 기본정보 4열 고정
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

        # (h) 오토필터 적용 후 '오토핏 근사 v2' 열 너비 설정
        #  - 한글/언더스코어 폭 가중치 + 필터 아이콘 여백 반영
        #  - 컬럼별 최소 폭 규칙 적용
        import re

        def east_asian_factor(s: str) -> float:
            """한글/全角 문자가 많을수록 가로폭이 더 필요하다는 가중치"""
            if not s:
                return 1.0
            # 한글/한자/전각류 포함 비율
            wide = len(re.findall(r'[\u1100-\u11FF\u3130-\u318F\uAC00-\uD7A3\u3000-\u303F\u3040-\u30FF\u4E00-\u9FFF]', s))
            ratio = wide / max(len(s), 1)
            # 기본 1.0 ~ 최대 1.25 가중
            return 1.0 + 0.25 * ratio

        # 숫자형/비율/차이 류 키워드
        numeric_like = {'순위','가격','원가','할인','평점','리뷰','차이','율','%','%p'}
        max_sample_rows = min(ws.max_row, 500)

        # 컬럼별 최소 폭(문자 수 기준, Excel width 단위 근사)
        min_widths = {
            # 기본 정보
            '카테고리': 20,
            '로켓_순위': 16,
            'UPC': 20,
            '품번': 20,

            # 로켓직구
            '로켓_상품ID': 20,
            '로켓_제품명': 56,
            '로켓_가격': 18,
            '로켓_원가': 18,
            '로켓_할인율(%)': 16,
            '로켓_평점': 16,
            '로켓_리뷰수': 16,
            '로켓_링크': 14,

            # 아이허브 공식
            '아이허브_상품ID': 20,
            '아이허브_카테고리': 24,
            '아이허브_제품명': 56,
            '아이허브_가격': 18,
            '아이허브_원가': 18,
            '아이허브_할인율(%)': 16,
            '아이허브_순위': 16,
            '아이허브_평점': 16,
            '아이허브_리뷰수': 16,
            '아이허브_링크': 14,

            # 분석
            '가격차이(원)': 18,
            '가격차이율(%)': 16,
            '더_저렴한_곳': 16,
            '할인율차이(%p)': 18,
            '평점차이': 16,
            '리뷰수차이': 16,
        }

        for col_idx, col_name in enumerate(header_row_values, start=1):
            col_letter = get_column_letter(col_idx)

            # 1) 헤더 길이(필터 아이콘 여백 + 한글 가중치)
            header_base = len(str(col_name))
            header_est  = int((header_base + 3) * east_asian_factor(str(col_name)))  # +3: 필터 아이콘/패딩

            # 2) 데이터 길이(최대 500행 샘플, 링크열은 "Link"로 평가)
            data_max_len = 0
            for r in range(3, max_sample_rows + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is None:
                    continue
                s = "Link" if '링크' in col_name else str(v)
                data_max_len = max(data_max_len, int(len(s) * east_asian_factor(s)))

            # 3) 기본 폭(헤더/데이터 중 큰 값에 10% 버퍼)
            est = int(max(header_est, data_max_len) * 1.10)

            # 4) 카테고리/제품명/링크/숫자형 등 특례 최소치 적용
            #    (정의된 최소 폭 우선, 없으면 분류 규칙)
            if col_name in min_widths:
                est = max(est, min_widths[col_name])
            else:
                if '제품명' in col_name:
                    est = max(est, 45)
                    est = min(est, 64)
                elif '링크' in col_name:
                    est = max(est, 10)
                    est = min(est, 14)
                elif any(key in col_name for key in numeric_like):
                    est = max(est, 12)
                    est = min(est, 18)
                else:
                    # 일반 텍스트는 12~32 사이
                    est = min(max(est, 12), 32)

            ws.column_dimensions[col_letter].width = est

    wb.save(output_path)
    print(f"\n✅ 엑셀 저장 완료: {output_path}")
    print(f"   📑 시트 개수: {len(wb.sheetnames)}개")
    print(f"   📅 포함 날짜: {', '.join(wb.sheetnames)}")


def generate_summary_stats(date_data_dict):
    """요약 통계 출력"""
    
    if not date_data_dict:
        return
    
    print("\n" + "="*80)
    print("📊 요약 통계")
    print("="*80)
    
    for date_str in sorted(date_data_dict.keys(), reverse=True):
        df = date_data_dict[date_str]
        
        if df.empty:
            continue
        
        print(f"\n📅 {date_str}")
        print("-" * 80)
        
        print(f"1️⃣ 전체 현황")
        print(f"   - 총 로켓직구 상품: {len(df):,}개")
        print(f"   - 매칭된 아이허브 상품: {df['iherb_vendor_id'].notna().sum():,}개")
        print(f"   - 매칭률: {df['iherb_vendor_id'].notna().sum() / len(df) * 100:.1f}%")
        
        # 매칭된 제품만 필터링
        matched_df = df[df['iherb_vendor_id'].notna()].copy()
        
        if not matched_df.empty:
            print(f"\n2️⃣ 가격 경쟁력 (매칭된 제품 기준)")
            cheaper_counts = matched_df['cheaper_source'].value_counts()
            for source, count in cheaper_counts.items():
                pct = count / len(matched_df) * 100
                print(f"   - {source}: {count:,}개 ({pct:.1f}%)")
            
            print(f"\n3️⃣ 가격 차이 통계 (매칭된 제품 기준)")
            print(f"   - 평균 가격차이: {matched_df['price_diff'].mean():,.0f}원")
            print(f"   - 중앙값: {matched_df['price_diff'].median():,.0f}원")
            print(f"   - 최대(아이허브가 비쌈): {matched_df['price_diff'].max():,.0f}원")
            print(f"   - 최소(로켓직구가 비쌈): {matched_df['price_diff'].min():,.0f}원")
            
            print(f"\n4️⃣ 카테고리별 매칭률")
            for category in df['category'].unique():
                cat_total = (df['category'] == category).sum()
                cat_matched = ((df['category'] == category) & (df['iherb_vendor_id'].notna())).sum()
                pct = cat_matched / cat_total * 100
                print(f"   - {category}: {pct:.1f}% ({cat_matched}/{cat_total})")
            
            print(f"\n5️⃣ 주목할 제품 (로켓직구 상위 20위 내)")
            top_rocket = matched_df[matched_df['rocket_rank'] <= 20].copy()
            if not top_rocket.empty:
                print(f"   - 총 {len(top_rocket)}개")
                top_rocket_iherb_cheaper = (top_rocket['cheaper_source'] == '아이허브').sum()
                print(f"   - 아이허브가 저렴: {top_rocket_iherb_cheaper}개")
                print(f"   - 로켓직구가 저렴: {len(top_rocket) - top_rocket_iherb_cheaper}개")
    
    print("="*80 + "\n")


def main():
    """메인 실행"""
    
    print("\n" + "="*80)
    print("🎯 로켓직구 전체 상품 기준 가격 비교 테이블 생성")
    print("="*80)
    print(f"DB 경로: {DB_PATH}")
    print(f"출력 경로: {OUTPUT_DIR}")
    print("="*80 + "\n")
    
    # 출력 디렉토리 생성
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # 사용 가능한 날짜 조회
    available_dates = get_available_dates()
    
    if not available_dates:
        print("❌ 사용 가능한 데이터가 없습니다.")
        return
    
    print(f"📅 사용 가능한 날짜: {len(available_dates)}개")
    for date in available_dates[:5]:  # 최근 5개만 표시
        print(f"   - {date}")
    if len(available_dates) > 5:
        print(f"   ... 외 {len(available_dates) - 5}개")
    
    dates_to_process = available_dates
    print(f"\n✅ 처리할 날짜: {', '.join(dates_to_process)}")
    
    date_data_dict = {}
    
    for target_date in dates_to_process:
        print(f"\n{'='*80}")
        df = extract_price_comparison_data(target_date=target_date)
        
        if not df.empty:
            date_data_dict[target_date] = df
    
    if not date_data_dict:
        print("\n❌ 처리할 데이터가 없습니다.")
        return
    
    # 요약 통계
    generate_summary_stats(date_data_dict)
    
    # 엑셀 리포트 생성
    today = datetime.now().strftime('%Y%m%d')
    output_path = os.path.join(OUTPUT_DIR, f"rocket_price_comparison_{today}.xlsx")
    create_excel_report(date_data_dict, output_path)
    
    print("\n✅ 처리 완료!")
    print(f"📁 파일 위치: {output_path}")
    print("\n💡 사용 팁:")
    print("   1. 각 시트는 날짜별로 구성되어 있습니다")
    print("   2. 로켓_순위 컬럼으로 정렬하여 상위 제품 확인")
    print("   3. 카테고리 필터로 특정 카테고리만 보기")
    print("   4. 매칭된 제품만 보려면 '아이허브_상품ID' 필터에서 공백 제외")
    print("   5. 링크는 'Link' 텍스트를 클릭하면 해당 페이지로 이동")
    print("   6. 조건부 서식: 초록=아이허브 저렴, 빨강=로켓직구 저렴")


if __name__ == "__main__":
    main()