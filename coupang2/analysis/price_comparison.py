#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
ê°€ê²© ë¹„êµ ë¦¬í¬íŠ¸ ìƒì„± (í†µí•© DB ë²„ì „)
â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
í†µí•© DBì—ì„œ ë°ì´í„°ë¥¼ ì½ì–´ Excel ë¦¬í¬íŠ¸ ìƒì„±
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from datetime import datetime
import sqlite3
import sys
from pathlib import Path

# í”„ë¡œì íŠ¸ ë£¨íŠ¸ ì¶”ê°€
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# í†µí•© ëª¨ë“ˆ import
from src.data_manager import DataManager


def get_available_dates(db_path):
    """ì‚¬ìš© ê°€ëŠ¥í•œ ë‚ ì§œ ëª©ë¡"""
    conn = sqlite3.connect(db_path)
    query = """
    SELECT DISTINCT snapshot_date as date
    FROM snapshots
    ORDER BY date DESC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df['date'].tolist()


def extract_price_comparison_data(db_path, target_date=None):
    """ê°€ê²© ë¹„êµ ë°ì´í„° ì¶”ì¶œ (í†µí•© DB ë²„ì „)
    
    Args:
        db_path: í†µí•© DB ê²½ë¡œ
        target_date: ëŒ€ìƒ ë‚ ì§œ (Noneì´ë©´ ìµœì‹ )
    
    Returns:
        DataFrame
    """
    
    print(f"\n{'='*80}")
    print(f"ğŸ“… ê°€ê²© ë¹„êµ ë°ì´í„° ì¶”ì¶œ (í†µí•© DB)")
    print(f"{'='*80}")
    print(f"ì²˜ë¦¬ ë‚ ì§œ: {target_date or 'ìµœì‹ '}")

    # DataManager ì‚¬ìš©
    manager = DataManager(db_path=db_path)
    
    # ë¯¸ë§¤ì¹­ ìš°ìˆ˜ ìƒí’ˆ í¬í•¨
    df = manager.get_integrated_df(target_date=target_date, include_unmatched=True)
    
    if df.empty:
        print("âš ï¸ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return df

    # ë¹„ì¤‘(%) ê³„ì‚°
    def calculate_share(colname, outname):
        """ì „ì²´ í•©ê³„ ëŒ€ë¹„ ë¹„ì¤‘ ê³„ì‚° (ì •ìˆ˜)"""
        total = pd.to_numeric(df[colname], errors='coerce').fillna(0).sum()
        if total <= 0:
            df[outname] = np.nan
        else:
            df[outname] = (pd.to_numeric(df[colname], errors='coerce').fillna(0) / total * 100).round(0).astype('Int64')

    share_columns = [
        ('iherb_revenue', 'ë§¤ì¶œë¹„ì¤‘'),
        ('iherb_sales_quantity', 'íŒë§¤ëŸ‰ë¹„ì¤‘'),
    ]
    
    for src_col, out_col in share_columns:
        if src_col in df.columns:
            calculate_share(src_col, out_col)
    
    # ì •ìˆ˜ ë³€í™˜
    if 'iherb_item_winner_ratio' in df.columns:
        df['iherb_item_winner_ratio'] = pd.to_numeric(df['iherb_item_winner_ratio'], errors='coerce').fillna(0).round(0).astype('Int64')

    print(f"\nâœ… ì´ {len(df):,}ê°œ ìƒí’ˆ")
    print(f"   - ë¡œì¼“ ë§¤ì¹­: {(df['matching_status'] == 'ë¡œì¼“ë§¤ì¹­').sum():,}ê°œ")
    print(f"   - ë¯¸ë§¤ì¹­ ìš°ìˆ˜: {(df['matching_status'] == 'ë¯¸ë§¤ì¹­').sum():,}ê°œ")
    
    # ì‹ ë¢°ë„ ë¶„í¬ (ë§¤ì¹­ ìƒí’ˆë§Œ)
    matched_df = df[df['matching_status'] == 'ë¡œì¼“ë§¤ì¹­']
    if len(matched_df) > 0:
        conf_counts = matched_df['matching_confidence'].value_counts()
        print(f"\n   ğŸ“Š ë§¤ì¹­ ì‹ ë¢°ë„ ë¶„í¬:")
        for conf, count in conf_counts.items():
            pct = count / len(matched_df) * 100
            print(f"      â€¢ {conf}: {count:,}ê°œ ({pct:.1f}%)")
    
    return df


def create_excel_report(date_data_dict, output_path):
    """Excel ë¦¬í¬íŠ¸ ìƒì„± - ë‹¨ì¼ ì‹œíŠ¸ (ë§¤ì¹­ + ë¯¸ë§¤ì¹­ í†µí•©)
    
    Args:
        date_data_dict: {ë‚ ì§œ: DataFrame}
        output_path: ì¶œë ¥ íŒŒì¼ ê²½ë¡œ
    """

    if not date_data_dict:
        print("âŒ ë°ì´í„°ê°€ ì—†ì–´ ì—‘ì…€ ìƒì„±ì„ ê±´ë„ˆëœë‹ˆë‹¤.")
        return

    print(f"\n{'='*80}")
    print(f"ğŸ“Š Excel ë¦¬í¬íŠ¸ ìƒì„± (ë‹¨ì¼ ì‹œíŠ¸ í†µí•©)")
    print(f"{'='*80}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for date_str, df in date_data_dict.items():
            if df.empty:
                continue

            # ì»¬ëŸ¼ ì¬êµ¬ì„± (41ê°œ)
            output_df = pd.DataFrame()

            # ========================================
            # 1ï¸âƒ£ ê¸°ë³¸ ì •ë³´ (8ê°œ) - ë§¤ì¹­ ìƒíƒœ í¬í•¨
            # ========================================
            output_df['ë§¤ì¹­ìƒíƒœ'] = df.get('matching_status', np.nan)
            output_df['ì‹ ë¢°ë„'] = df.get('matching_confidence', np.nan).apply(
                lambda x: '' if pd.isna(x) or x == '' else x
            )
            output_df['ë¡œì¼“_ì¹´í…Œê³ ë¦¬'] = df.get('rocket_category', np.nan)
            output_df['ì•„ì´í—ˆë¸Œ_ì¹´í…Œê³ ë¦¬'] = df.get('iherb_category', np.nan)
            output_df['ë¡œì¼“_ë§í¬'] = df.get('rocket_url', np.nan)
            output_df['ì•„ì´í—ˆë¸Œ_ë§í¬'] = df.get('iherb_url', np.nan)
            output_df['í’ˆë²ˆ'] = df.get('iherb_part_number', np.nan)
            output_df['UPC'] = pd.to_numeric(df.get('iherb_upc', np.nan), errors='coerce').astype('Int64')

            # ========================================
            # 2ï¸âƒ£ í•µì‹¬ ì§€í‘œ (9ê°œ)
            # ========================================
            output_df['ìˆœìœ„'] = df.get('rocket_rank', np.nan)
            output_df['íŒë§¤ëŸ‰'] = df.get('iherb_sales_quantity', np.nan)
            output_df['ë§¤ì¶œ(ì›)'] = df.get('iherb_revenue', np.nan)
            output_df['ì•„ì´í…œìœ„ë„ˆë¹„ìœ¨'] = df.get('iherb_item_winner_ratio', np.nan)
            output_df['ê°€ê²©ê²©ì°¨(ì›)'] = df.get('price_diff', np.nan)
            output_df['ì†ìµë¶„ê¸°í• ì¸ìœ¨'] = df.get('breakeven_discount_rate', np.nan)
            output_df['ì¶”ì²œí• ì¸ìœ¨'] = df.get('recommended_discount_rate', np.nan)
            output_df['ìš”ì²­í• ì¸ìœ¨'] = df.get('requested_discount_rate', np.nan)
            output_df['ìœ ë¦¬í•œê³³'] = df.get('cheaper_source', np.nan)

            # ========================================
            # 3ï¸âƒ£ ì œí’ˆ ì •ë³´ (7ê°œ)
            # ========================================
            output_df['ë¡œì¼“_ì œí’ˆëª…'] = df.get('rocket_product_name', np.nan)
            output_df['ì•„ì´í—ˆë¸Œ_ì œí’ˆëª…'] = df.get('iherb_product_name', np.nan)
            output_df['Product_ID'] = df.get('rocket_product_id', np.nan)
            output_df['ë¡œì¼“_Vendor'] = df.get('rocket_vendor_id', np.nan)
            output_df['ë¡œì¼“_Item'] = df.get('rocket_item_id', np.nan)
            output_df['ì•„ì´í—ˆë¸Œ_Vendor'] = df.get('iherb_vendor_id', np.nan)
            output_df['ì•„ì´í—ˆë¸Œ_Item'] = df.get('iherb_item_id', np.nan)

            # ========================================
            # 4ï¸âƒ£ ê°€ê²© ì •ë³´ (8ê°œ)
            # ========================================
            output_df['ì •ê°€'] = df.get('rocket_original_price', np.nan)
            output_df['í• ì¸ìœ¨'] = df.get('rocket_discount_rate', np.nan)
            output_df['ë¡œì¼“ê°€ê²©'] = df.get('rocket_price', np.nan)
            output_df['íŒë§¤ê°€'] = df.get('iherb_price', np.nan)
            output_df['ì •ê°€_ì•„ì´í—ˆë¸Œ'] = df.get('iherb_original_price', np.nan)
            output_df['ì¿ íŒ¡ì¶”ì²œê°€'] = df.get('iherb_recommended_price', np.nan)
            output_df['ì¬ê³ '] = df.get('iherb_stock', np.nan)
            output_df['íŒë§¤ìƒíƒœ'] = df.get('iherb_stock_status', np.nan)

            # ========================================
            # 5ï¸âƒ£ íŒë§¤ ì„±ê³¼ (7ê°œ) - ê°„ì†Œí™”
            # ========================================
            output_df['í‰ì '] = df.get('rocket_rating', np.nan)
            output_df['ë¦¬ë·°ìˆ˜'] = df.get('rocket_reviews', np.nan)
            output_df['ë§¤ì¶œë¹„ì¤‘'] = df.get('ë§¤ì¶œë¹„ì¤‘', np.nan)
            output_df['ì£¼ë¬¸'] = np.nan  # í†µí•© DBì— ì—†ìŒ
            output_df['íŒë§¤ëŸ‰ë¹„ì¤‘'] = df.get('íŒë§¤ëŸ‰ë¹„ì¤‘', np.nan)
            output_df['êµ¬ë§¤ì „í™˜ìœ¨'] = np.nan  # í†µí•© DBì— ì—†ìŒ
            output_df['ì·¨ì†Œìœ¨'] = np.nan  # í†µí•© DBì— ì—†ìŒ

            # ì‹œíŠ¸ ì‘ì„± (í—¤ë” ì—†ì´)
            sheet_name = date_str.replace('-', '')[:10]  # YYYYMMDD
            output_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            print(f"   âœ“ ì‹œíŠ¸ '{sheet_name}' ì‘ì„± ì™„ë£Œ ({len(output_df):,}ê°œ)")

    # ìŠ¤íƒ€ì¼ ì ìš©
    apply_excel_styles(output_path)

    print(f"\nâœ… Excel ë¦¬í¬íŠ¸ ìƒì„± ì™„ë£Œ: {output_path}")


def apply_excel_styles(output_path):
    """Excel ìŠ¤íƒ€ì¼ ì ìš© - 3ë‹¨ í—¤ë” + ë§¤ì¹­ìƒíƒœ êµ¬ë¶„"""

    wb = load_workbook(output_path)

    # ìƒ‰ìƒ íŒ”ë ˆíŠ¸
    INFO_DARK  = "0F172A"
    INFO_MID   = "475569"
    INFO_LIGHT = "E2E8F0"
    PRIMARY_DARK = "5E2A8A"
    PRIMARY_MID  = "7A3EB1"
    PRIMARY_LIGHT= "D2B7E5"
    SECONDARY_DARK = "305496"
    SECONDARY_MID = "4472C4"
    SECONDARY_LIGHT = "B4C7E7"
    TERTIARY_DARK = "C55A11"
    TERTIARY_MID = "F4B084"
    TERTIARY_LIGHT = "FBE5D6"
    SUCCESS_DARK = "375623"
    SUCCESS_MID = "548235"
    SUCCESS_LIGHT = "A8D08D"
    
    HIGHLIGHT_GREEN = "C6EFCE"
    HIGHLIGHT_RED = "FFC7CE"
    HIGHLIGHT_YELLOW = "FFEB9C"

    header_font_white = Font(color="FFFFFF", bold=True, size=11)
    header_font_dark = Font(color="000000", bold=True, size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ì»¬ëŸ¼ ê·¸ë£¹ ì •ì˜ (39ê°œ)
    column_groups = [
        # 1ï¸âƒ£ ê¸°ë³¸ ì •ë³´ (8ê°œ)
        {
            'name': 'ê¸°ë³¸ ì •ë³´',
            'color_top': INFO_DARK,
            'color_mid': INFO_MID,
            'color_bottom': INFO_LIGHT,
            'sub_groups': [
                {'name': 'ë§¤ì¹­', 'cols': ['ìƒíƒœ', 'ì‹ ë¢°ë„']},
                {'name': 'ì¹´í…Œê³ ë¦¬', 'cols': ['ë¡œì¼“', 'ì•„ì´í—ˆë¸Œ']},
                {'name': 'ë§í¬', 'cols': ['ë¡œì¼“', 'ì•„ì´í—ˆë¸Œ']},
                {'name': 'ìƒí’ˆë²ˆí˜¸', 'cols': ['í’ˆë²ˆ', 'UPC']}
            ]
        },
        # 2ï¸âƒ£ í•µì‹¬ ì§€í‘œ (9ê°œ)
        {
            'name': 'í•µì‹¬ ì§€í‘œ',
            'color_top': PRIMARY_DARK,
            'color_mid': PRIMARY_MID,
            'color_bottom': PRIMARY_LIGHT,
            'sub_groups': [
                {'name': 'ë¡œì¼“', 'cols': ['ìˆœìœ„']},
                {'name': 'ì•„ì´í—ˆë¸Œ', 'cols': ['íŒë§¤ëŸ‰', 'ë§¤ì¶œ(ì›)', 'ì•„ì´í…œìœ„ë„ˆë¹„ìœ¨']},
                {'name': 'ì¢…í•©', 'cols': ['ê°€ê²©ê²©ì°¨(ì›)', 'ì†ìµë¶„ê¸°í• ì¸ìœ¨', 'ì¶”ì²œí• ì¸ìœ¨', 'ìš”ì²­í• ì¸ìœ¨', 'ìœ ë¦¬í•œê³³']}
            ]
        },
        # 3ï¸âƒ£ ì œí’ˆ ì •ë³´ (7ê°œ)
        {
            'name': 'ì œí’ˆ ì •ë³´',
            'color_top': SECONDARY_DARK,
            'color_mid': SECONDARY_MID,
            'color_bottom': SECONDARY_LIGHT,
            'sub_groups': [
                {'name': 'ì œí’ˆëª…', 'cols': ['ë¡œì¼“', 'ì•„ì´í—ˆë¸Œ']},
                {
                    'name': 'ìƒí’ˆ ID', 
                    'cols': ['Product_ID', 'ë¡œì¼“_Vendor', 'ë¡œì¼“_Item', 'ì•„ì´í—ˆë¸Œ_Vendor', 'ì•„ì´í—ˆë¸Œ_Item']
                }
            ]
        },
        # 4ï¸âƒ£ ê°€ê²© ì •ë³´ (8ê°œ)
        {
            'name': 'ê°€ê²© ì •ë³´',
            'color_top': TERTIARY_DARK,
            'color_mid': TERTIARY_MID,
            'color_bottom': TERTIARY_LIGHT,
            'sub_groups': [
                {'name': 'ë¡œì¼“ì§êµ¬', 'cols': ['ì •ê°€', 'í• ì¸ìœ¨', 'ë¡œì¼“ê°€ê²©']},
                {'name': 'ì•„ì´í—ˆë¸Œ', 'cols': ['íŒë§¤ê°€', 'ì •ê°€', 'ì¿ íŒ¡ì¶”ì²œê°€', 'ì¬ê³ ', 'íŒë§¤ìƒíƒœ']}
            ]
        },
        # 5ï¸âƒ£ íŒë§¤ ì„±ê³¼ (7ê°œ)
        {
            'name': 'íŒë§¤ ì„±ê³¼',
            'color_top': SUCCESS_DARK,
            'color_mid': SUCCESS_MID,
            'color_bottom': SUCCESS_LIGHT,
            'sub_groups': [
                {'name': 'ë¡œì¼“', 'cols': ['í‰ì ', 'ë¦¬ë·°ìˆ˜']},
                {'name': 'ì•„ì´í—ˆë¸Œ', 'cols': ['ë§¤ì¶œë¹„ì¤‘', 'ì£¼ë¬¸', 'íŒë§¤ëŸ‰ë¹„ì¤‘', 'êµ¬ë§¤ì „í™˜ìœ¨', 'ì·¨ì†Œìœ¨']}
            ]
        }
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 3ê°œ í–‰ ì‚½ì…
        ws.insert_rows(1, 3)

        # í—¤ë” ì‘ì„±
        col_pos = 1
        
        for group in column_groups:
            group_name = group['name']
            color_top = group['color_top']
            color_mid = group['color_mid']
            color_bottom = group['color_bottom']
            sub_groups = group['sub_groups']
            
            total_span = sum(len(sg['cols']) for sg in sub_groups)
            
            # ìƒìœ„ í—¤ë”
            ws.merge_cells(start_row=1, start_column=col_pos,
                         end_row=1, end_column=col_pos + total_span - 1)
            cell_top = ws.cell(row=1, column=col_pos)
            cell_top.value = group_name
            cell_top.fill = PatternFill(start_color=color_top, end_color=color_top, fill_type="solid")
            cell_top.font = header_font_white
            cell_top.alignment = Alignment(horizontal='center', vertical='center')
            cell_top.border = thin_border
            
            # ì¤‘ê°„ í—¤ë”
            for sub_group in sub_groups:
                sub_name = sub_group['name']
                sub_cols = sub_group['cols']
                sub_span = len(sub_cols)
                
                ws.merge_cells(start_row=2, start_column=col_pos,
                             end_row=2, end_column=col_pos + sub_span - 1)
                cell_mid = ws.cell(row=2, column=col_pos)
                cell_mid.value = sub_name
                cell_mid.fill = PatternFill(start_color=color_mid, end_color=color_mid, fill_type="solid")
                cell_mid.font = header_font_white
                cell_mid.alignment = Alignment(horizontal='center', vertical='center')
                cell_mid.border = thin_border
                
                # í•˜ìœ„ í—¤ë”
                for i, col_name in enumerate(sub_cols):
                    cell_bottom = ws.cell(row=3, column=col_pos + i)
                    cell_bottom.value = col_name
                    cell_bottom.fill = PatternFill(start_color=color_bottom, end_color=color_bottom, fill_type="solid")
                    cell_bottom.font = header_font_dark
                    cell_bottom.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell_bottom.border = thin_border
                
                col_pos += sub_span

        # ì»¬ëŸ¼ ë„ˆë¹„ ì„¤ì •
        header_names = [cell.value for cell in ws[3]]
        
        def col_idx_of(name):
            try:
                return header_names.index(name) + 1
            except ValueError:
                return None
        
        column_widths = {
            'ìƒíƒœ': 7.9,
            'ì‹ ë¢°ë„': 9.71,
            'ë¡œì¼“': 12.86,
            'ì•„ì´í—ˆë¸Œ': 11.00,
            'í’ˆë²ˆ': 12.71,
            'UPC': 15.00,
            'ìˆœìœ„': 7.86,
            'íŒë§¤ëŸ‰': 9.00,
            'ë§¤ì¶œ(ì›)': 10.14,
            'ì•„ì´í…œìœ„ë„ˆë¹„ìœ¨': 15.57,
            'ê°€ê²©ê²©ì°¨(ì›)': 13.43,
            'ì†ìµë¶„ê¸°í• ì¸ìœ¨': 15.57,
            'ì¶”ì²œí• ì¸ìœ¨': 12.29,
            'ìš”ì²­í• ì¸ìœ¨': 12.29,
            'ìœ ë¦¬í•œê³³': 10.57,
            'Product_ID': 14.14,
            'ë¡œì¼“_Vendor': 17.29,
            'ë¡œì¼“_Item': 15.14,
            'ì•„ì´í—ˆë¸Œ_Vendor': 17.29,
            'ì•„ì´í—ˆë¸Œ_Item': 15.14,
            'ì •ê°€': 8.86,
            'í• ì¸ìœ¨': 9.00,
            'ë¡œì¼“ê°€ê²©': 10.57,
            'íŒë§¤ê°€': 10.57,
            'ì¿ íŒ¡ì¶”ì²œê°€': 12.29,
            'ì¬ê³ ': 7.29,
            'íŒë§¤ìƒíƒœ': 10.57,
            'í‰ì ': 8.86,
            'ë¦¬ë·°ìˆ˜': 9.00,
            'ë§¤ì¶œë¹„ì¤‘': 10.57,
            'ì£¼ë¬¸': 7.29,
            'íŒë§¤ëŸ‰ë¹„ì¤‘': 12.29,
            'êµ¬ë§¤ì „í™˜ìœ¨': 13.00,
            'ì·¨ì†Œìœ¨': 9.00,
        }

        DEFAULT_WIDTH = 12
        PRODUCT_NAME_WIDTH = 60.0

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            
            mid_header = ws.cell(row=2, column=col_idx).value
            if mid_header is None and col_idx > 1:
                mid_header = ws.cell(row=2, column=col_idx - 1).value

            bottom_header = ws.cell(row=3, column=col_idx).value

            if mid_header == 'ì œí’ˆëª…':
                width = PRODUCT_NAME_WIDTH
            elif mid_header == 'ë§í¬' and bottom_header == 'ë¡œì¼“':
                width = 7.29
            elif mid_header == 'ë§í¬' and bottom_header == 'ì•„ì´í—ˆë¸Œ':
                width = 10.57
            elif bottom_header in column_widths:
                width = column_widths[bottom_header]
            else:
                width = DEFAULT_WIDTH

            ws.column_dimensions[col_letter].width = width
        
        # ë°ì´í„° ì…€ ê¸°ë³¸ ìŠ¤íƒ€ì¼
        data_actual_start = 4
        max_col = ws.max_column
        
        for row_idx in range(data_actual_start, ws.max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=False)

        # ì¡°ê±´ë¶€ ì„œì‹
        winner_col = col_idx_of('ì•„ì´í…œìœ„ë„ˆë¹„ìœ¨')
        if winner_col:
            for row_idx in range(data_actual_start, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=winner_col)
                try:
                    val = float(cell.value) if cell.value else 0
                    if val >= 30:
                        cell.fill = PatternFill(start_color=HIGHLIGHT_GREEN, end_color=HIGHLIGHT_GREEN, fill_type="solid")
                except:
                    pass
        
        breakeven_col = col_idx_of('ì†ìµë¶„ê¸°í• ì¸ìœ¨')
        if breakeven_col:
            for row_idx in range(data_actual_start, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=breakeven_col)
                try:
                    val = float(cell.value) if cell.value else 0
                    if val > 0:
                        cell.fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
                except:
                    pass
        
        recommended_col = col_idx_of('ì¶”ì²œí• ì¸ìœ¨')
        if recommended_col:
            for row_idx in range(data_actual_start, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=recommended_col)
                try:
                    val = float(cell.value) if cell.value else 0
                    if val > 0:
                        cell.fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
                except:
                    pass
        
        requested_col = col_idx_of('ìš”ì²­í• ì¸ìœ¨')
        if requested_col:
            for row_idx in range(data_actual_start, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=requested_col)
                try:
                    val = float(cell.value) if cell.value else 0
                    if val > 0:
                        cell.fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
                except:
                    pass
        
        price_diff_col = col_idx_of('ê°€ê²©ê²©ì°¨(ì›)')
        cheaper_col = col_idx_of('ìœ ë¦¬í•œê³³')
        if price_diff_col and cheaper_col:
            for row_idx in range(data_actual_start, ws.max_row + 1):
                cheaper_value = ws.cell(row=row_idx, column=cheaper_col).value
                if cheaper_value == 'ì•„ì´í—ˆë¸Œ':
                    ws.cell(row=row_idx, column=price_diff_col).fill = PatternFill(start_color=HIGHLIGHT_GREEN, end_color=HIGHLIGHT_GREEN, fill_type="solid")
                elif cheaper_value == 'ë¡œì¼“ì§êµ¬':
                    ws.cell(row=row_idx, column=price_diff_col).fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
        
        conf_col = col_idx_of('ì‹ ë¢°ë„')
        if conf_col:
            for row_idx in range(data_actual_start, ws.max_row + 1):
                conf_value = ws.cell(row=row_idx, column=conf_col).value
                cell = ws.cell(row=row_idx, column=conf_col)
                if conf_value == 'High':
                    cell.fill = PatternFill(start_color=HIGHLIGHT_GREEN, end_color=HIGHLIGHT_GREEN, fill_type="solid")
                elif conf_value == 'Medium':
                    cell.fill = PatternFill(start_color=HIGHLIGHT_YELLOW, end_color=HIGHLIGHT_YELLOW, fill_type="solid")
                elif conf_value == 'Low':
                    cell.fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")

        # í•˜ì´í¼ë§í¬
        link_columns = []
        for col_idx in range(1, max_col + 1):
            mid_header = ws.cell(row=2, column=col_idx).value
            
            if mid_header == 'ë§í¬':
                link_columns.append(col_idx)
            elif mid_header is None and col_idx > 1:
                left_mid_header = ws.cell(row=2, column=col_idx - 1).value
                if left_mid_header == 'ë§í¬':
                    link_columns.append(col_idx)
        
        for col_idx in link_columns:
            for row_idx in range(data_actual_start, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                url = cell.value
                if url and str(url).strip() and str(url).startswith('http'):
                    cell.value = "Link"
                    cell.hyperlink = str(url)
                    cell.font = Font(color="0563C1", underline="single")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Freeze Panes
        freeze_col = 17
        ws.freeze_panes = ws.cell(row=4, column=freeze_col)

        # ë°ì´í„°ë°”
        share_cols = ['ë§¤ì¶œë¹„ì¤‘', 'íŒë§¤ëŸ‰ë¹„ì¤‘']
        for share_col_name in share_cols:
            share_col_idx = col_idx_of(share_col_name)
            if share_col_idx:
                col_letter = get_column_letter(share_col_idx)
                rule = DataBarRule(
                    start_type='num', start_value=0,
                    end_type='num', end_value=100,
                    color="63C384"
                )
                ws.conditional_formatting.add(
                    f'{col_letter}{data_actual_start}:{col_letter}{ws.max_row}',
                    rule
                )
        
        # ìë™ í•„í„°
        ws.auto_filter.ref = (
            f"{get_column_letter(1)}{3}:"
            f"{get_column_letter(ws.max_column)}{ws.max_row}"
        )

    wb.save(output_path)


def main():
    """ë©”ì¸ í•¨ìˆ˜"""
    
    # Config import
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from config import Config
    
    DB_PATH = Config.INTEGRATED_DB_PATH
    OUTPUT_DIR = Config.OUTPUT_DIR
    
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    dates = get_available_dates(DB_PATH)
    
    if not dates:
        print("âŒ ì‚¬ìš© ê°€ëŠ¥í•œ ë‚ ì§œê°€ ì—†ìŠµë‹ˆë‹¤.")
        return
    
    print(f"\nì‚¬ìš© ê°€ëŠ¥í•œ ë‚ ì§œ: {len(dates)}ê°œ")
    for i, date in enumerate(dates[:5], 1):
        print(f"  {i}. {date}")
    
    process_dates = dates[:1]
    
    date_data_dict = {}
    for date_str in process_dates:
        df = extract_price_comparison_data(DB_PATH, target_date=date_str)
        if not df.empty:
            date_data_dict[date_str] = df
    
    if date_data_dict:
        output_file = OUTPUT_DIR / f"rocket_vs_iherb_{datetime.now().strftime('%Y%m%d')}.xlsx"
        create_excel_report(date_data_dict, str(output_file))
    else:
        print("\nâŒ ìƒì„±í•  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")


if __name__ == "__main__":
    main()