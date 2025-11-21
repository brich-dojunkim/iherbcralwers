#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
상품별 추이 대시보드
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
전일 대비 변화 추이를 보여주는 상품 레벨 대시보드
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sqlite3
import sys
from pathlib import Path

# 프로젝트 루트 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from src.data_manager import DataManager


def get_trend_data(db_path: str) -> pd.DataFrame:
    """전일 대비 추이 데이터 생성
    
    Returns:
        DataFrame with columns:
        - 기본 정보 (품번, UPC, 카테고리, 링크 등)
        - 현재 상태 (판매량, 재고, 가격, 순위)
        - 전일 대비 변화 (판매량↔, 재고↔, 가격↔, 순위↔)
    """
    
    print(f"\n{'='*80}")
    print(f"📊 상품별 추이 대시보드 데이터 생성")
    print(f"{'='*80}\n")
    
    conn = sqlite3.connect(db_path)
    
    # 최신 2개 스냅샷 확인
    cursor = conn.execute("""
        SELECT id, snapshot_date 
        FROM snapshots 
        ORDER BY snapshot_date DESC, id DESC 
        LIMIT 2
    """)
    snapshots = cursor.fetchall()
    
    if len(snapshots) < 2:
        print("❌ 비교를 위한 스냅샷이 부족합니다 (최소 2개 필요)")
        conn.close()
        return pd.DataFrame()
    
    current_id, current_date = snapshots[0]
    prev_id, prev_date = snapshots[1]
    
    print(f"📅 비교 기준:")
    print(f"   현재: Snapshot {current_id} ({current_date})")
    print(f"   전일: Snapshot {prev_id} ({prev_date})")
    print()
    
    # 현재/전일 스냅샷 데이터
    manager = DataManager(db_path)
    df_current = manager.get_integrated_df(snapshot_id=current_id, include_unmatched=False)
    df_prev = manager.get_integrated_df(snapshot_id=prev_id, include_unmatched=False)
    
    if df_current.empty:
        print("❌ 현재 스냅샷 데이터가 없습니다")
        conn.close()
        return pd.DataFrame()
    
    print(f"✅ 데이터 로드:")
    print(f"   현재: {len(df_current):,}개")
    print(f"   전일: {len(df_prev):,}개\n")
    
    # 아이허브 순위 계산 (판매량 기준)
    df_current['iherb_rank_current'] = df_current['iherb_sales_quantity'].rank(
        method='min', ascending=False, na_option='bottom'
    ).astype('Int64')
    
    df_prev['iherb_rank_prev'] = df_prev['iherb_sales_quantity'].rank(
        method='min', ascending=False, na_option='bottom'
    ).astype('Int64')
    
    # 전일 데이터에서 필요한 컬럼만 추출
    prev_cols = {
        'iherb_sales_quantity': 'sales_prev',
        'iherb_stock': 'stock_prev',
        'iherb_price': 'price_prev',
        'rocket_price': 'rocket_price_prev',
        'rocket_rank': 'rocket_rank_prev',
        'iherb_rank_prev': 'iherb_rank_prev'
    }
    
    df_prev_subset = df_prev[['iherb_vendor_id'] + list(prev_cols.keys())].copy()
    df_prev_subset = df_prev_subset.rename(columns=prev_cols)
    
    # ============================
    # 🔑 iherb_vendor_id 기준 안전 머지
    #   - 키 있는 행만 전일/당일 비교
    #   - 키 없는 행은 전일값 NaN으로 처리
    # ============================
    
    # 1) 키 있는 행들만 분리
    curr_has_key = df_current['iherb_vendor_id'].notna()
    prev_has_key = df_prev_subset['iherb_vendor_id'].notna()
    
    curr_key = df_current[curr_has_key].copy()
    prev_key = df_prev_subset[prev_has_key].copy()
    
    # 전일 데이터: 동일 iherb_vendor_id 여러 줄 있을 수 있으니 1개만 남김
    prev_key = prev_key.drop_duplicates(subset='iherb_vendor_id', keep='last')
    
    # 키 있는 행들만 merge
    df_merged_key = curr_key.merge(
        prev_key,
        on='iherb_vendor_id',
        how='left'
    )
    
    # 2) 키 없는 행들: 전일값 전부 NaN으로 세팅
    curr_nokey = df_current[~curr_has_key].copy()
    for col in ['sales_prev', 'stock_prev', 'price_prev',
                'rocket_price_prev', 'rocket_rank_prev', 'iherb_rank_prev']:
        curr_nokey[col] = np.nan
    
    # 3) 다시 합치기
    df_merged = pd.concat([df_merged_key, curr_nokey], ignore_index=True)
    
    # 변화량 계산
    df_merged['sales_change'] = df_merged['iherb_sales_quantity'] - df_merged['sales_prev']
    df_merged['stock_change'] = df_merged['iherb_stock'] - df_merged['stock_prev']
    df_merged['price_change'] = df_merged['iherb_price'] - df_merged['price_prev']
    df_merged['rocket_price_change'] = df_merged['rocket_price'] - df_merged['rocket_price_prev']
    
    # 순위 변화 (양수면 상승)
    df_merged['rocket_rank_change'] = df_merged['rocket_rank_prev'] - df_merged['rocket_rank']
    df_merged['iherb_rank_change'] = df_merged['iherb_rank_prev'] - df_merged['iherb_rank_current']
    
    # 재고 소진 예상일
    df_merged['days_to_stockout'] = np.where(
        df_merged['iherb_sales_quantity'] > 0,
        df_merged['iherb_stock'] / df_merged['iherb_sales_quantity'],
        np.nan
    )
    
    # 통계
    print(f"📈 변화 통계:")
    print(f"   판매량 증가: {(df_merged['sales_change'] > 0).sum():,}개")
    print(f"   판매량 감소: {(df_merged['sales_change'] < 0).sum():,}개")
    print(f"   판매량 동일: {(df_merged['sales_change'] == 0).sum():,}개")
    print(f"   신규 상품: {df_merged['sales_prev'].isna().sum():,}개\n")
    
    conn.close()
    return df_merged


def create_dashboard_excel(df: pd.DataFrame, output_path: str):
    """대시보드 Excel 생성
    
    Args:
        df: 추이 데이터
        output_path: 출력 경로
    """
    
    if df.empty:
        print("❌ 데이터가 없어 Excel 생성을 건너뜁니다.")
        return
    
    print(f"📊 Excel 대시보드 생성 중...\n")
    
    # 컬럼 구성
    output_df = pd.DataFrame()
    
    # ========================================
    # 1️⃣ 기본 정보 (8개)
    # ========================================
    output_df['매칭상태'] = df.get('matching_status', np.nan)
    output_df['신뢰도'] = df.get('matching_confidence', np.nan)
    output_df['로켓_카테고리'] = df.get('rocket_category', np.nan)
    output_df['아이허브_카테고리'] = df.get('iherb_category', np.nan)
    output_df['로켓_링크'] = df.get('rocket_url', np.nan)
    output_df['아이허브_링크'] = df.get('iherb_url', np.nan)
    output_df['품번'] = df.get('iherb_part_number', np.nan)
    output_df['UPC'] = pd.to_numeric(df.get('iherb_upc', np.nan), errors='coerce').astype('Int64')
    
    # ========================================
    # 2️⃣ 현재 상태 (9개)
    # ========================================
    output_df['판매량'] = df.get('iherb_sales_quantity', np.nan)
    output_df['재고'] = df.get('iherb_stock', np.nan)
    output_df['재고소진일'] = pd.to_numeric(df.get('days_to_stockout', np.nan), errors='coerce').round(0).astype('Int64')
    output_df['판매가'] = df.get('iherb_price', np.nan)
    output_df['로켓가격'] = df.get('rocket_price', np.nan)
    output_df['로켓순위'] = df.get('rocket_rank', np.nan)
    output_df['아이허브순위'] = df.get('iherb_rank_current', np.nan)
    output_df['매출'] = df.get('iherb_revenue', np.nan)
    output_df['아이템위너비율'] = df.get('iherb_item_winner_ratio', np.nan)
    
    # ========================================
    # 3️⃣ 전일 대비 변화 (6개)
    # ========================================
    output_df['판매량↔'] = df.get('sales_change', np.nan).astype('Int64')
    output_df['재고↔'] = df.get('stock_change', np.nan).astype('Int64')
    output_df['판매가↔'] = df.get('price_change', np.nan).astype('Int64')
    output_df['로켓가격↔'] = df.get('rocket_price_change', np.nan).astype('Int64')
    output_df['로켓순위↔'] = df.get('rocket_rank_change', np.nan).astype('Int64')
    output_df['아이허브순위↔'] = df.get('iherb_rank_change', np.nan).astype('Int64')
    
    # ========================================
    # 4️⃣ 제품 정보 (7개)
    # ========================================
    output_df['로켓_제품명'] = df.get('rocket_product_name', np.nan)
    output_df['아이허브_제품명'] = df.get('iherb_product_name', np.nan)
    output_df['Product_ID'] = df.get('rocket_product_id', np.nan)
    output_df['로켓_Vendor'] = df.get('rocket_vendor_id', np.nan)
    output_df['로켓_Item'] = df.get('rocket_item_id', np.nan)
    output_df['아이허브_Vendor'] = df.get('iherb_vendor_id', np.nan)
    output_df['아이허브_Item'] = df.get('iherb_item_id', np.nan)
    
    # Excel 작성
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='상품별_추이', index=False, header=False)
    
    # 스타일 적용
    apply_dashboard_styles(output_path)
    
    print(f"✅ Excel 생성 완료: {output_path}\n")


def apply_dashboard_styles(output_path: str):
    """Excel 스타일 적용"""
    
    wb = load_workbook(output_path)
    ws = wb['상품별_추이']
    
    # 색상
    INFO_DARK = "0F172A"
    INFO_MID = "475569"
    INFO_LIGHT = "E2E8F0"
    
    STATUS_DARK = "7C3AED"
    STATUS_MID = "A78BFA"
    STATUS_LIGHT = "DDD6FE"
    
    TREND_DARK = "DC2626"
    TREND_MID = "F87171"
    TREND_LIGHT = "FEE2E2"
    
    PRODUCT_DARK = "0891B2"
    PRODUCT_MID = "22D3EE"
    PRODUCT_LIGHT = "CFFAFE"
    
    HIGHLIGHT_GREEN = "C6EFCE"
    HIGHLIGHT_RED = "FFC7CE"
    HIGHLIGHT_YELLOW = "FFEB9C"
    
    header_font_white = Font(color="FFFFFF", bold=True, size=11)
    header_font_dark = Font(color="000000", bold=True, size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 3단 헤더 삽입
    ws.insert_rows(1, 3)
    
    # 컬럼 그룹
    column_groups = [
        {
            'name': '기본 정보',
            'color_top': INFO_DARK,
            'color_mid': INFO_MID,
            'color_bottom': INFO_LIGHT,
            'sub_groups': [
                {'name': '매칭', 'cols': ['상태', '신뢰도']},
                {'name': '카테고리', 'cols': ['로켓', '아이허브']},
                {'name': '링크', 'cols': ['로켓', '아이허브']},
                {'name': '상품번호', 'cols': ['품번', 'UPC']}
            ]
        },
        {
            'name': '현재 상태',
            'color_top': STATUS_DARK,
            'color_mid': STATUS_MID,
            'color_bottom': STATUS_LIGHT,
            'sub_groups': [
                {'name': '판매', 'cols': ['판매량', '재고', '재고소진일']},
                {'name': '가격', 'cols': ['판매가', '로켓가격']},
                {'name': '순위', 'cols': ['로켓순위', '아이허브순위']},
                {'name': '성과', 'cols': ['매출', '아이템위너비율']}
            ]
        },
        {
            'name': '전일 대비',
            'color_top': TREND_DARK,
            'color_mid': TREND_MID,
            'color_bottom': TREND_LIGHT,
            'sub_groups': [
                {'name': '판매', 'cols': ['판매량↔', '재고↔']},
                {'name': '가격', 'cols': ['판매가↔', '로켓가격↔']},
                {'name': '순위', 'cols': ['로켓순위↔', '아이허브순위↔']}
            ]
        },
        {
            'name': '제품 정보',
            'color_top': PRODUCT_DARK,
            'color_mid': PRODUCT_MID,
            'color_bottom': PRODUCT_LIGHT,
            'sub_groups': [
                {'name': '제품명', 'cols': ['로켓', '아이허브']},
                {'name': '상품 ID', 'cols': ['Product_ID', '로켓_Vendor', '로켓_Item', '아이허브_Vendor', '아이허브_Item']}
            ]
        }
    ]
    
    # 헤더 작성
    col_pos = 1
    
    for group in column_groups:
        total_span = sum(len(sg['cols']) for sg in group['sub_groups'])
        
        # 상위 헤더
        ws.merge_cells(start_row=1, start_column=col_pos,
                      end_row=1, end_column=col_pos + total_span - 1)
        cell_top = ws.cell(row=1, column=col_pos)
        cell_top.value = group['name']
        cell_top.fill = PatternFill(start_color=group['color_top'], 
                                    end_color=group['color_top'], fill_type="solid")
        cell_top.font = header_font_white
        cell_top.alignment = Alignment(horizontal='center', vertical='center')
        cell_top.border = thin_border
        
        # 중간/하위 헤더
        for sub_group in group['sub_groups']:
            sub_span = len(sub_group['cols'])
            
            ws.merge_cells(start_row=2, start_column=col_pos,
                         end_row=2, end_column=col_pos + sub_span - 1)
            cell_mid = ws.cell(row=2, column=col_pos)
            cell_mid.value = sub_group['name']
            cell_mid.fill = PatternFill(start_color=group['color_mid'], 
                                       end_color=group['color_mid'], fill_type="solid")
            cell_mid.font = header_font_white
            cell_mid.alignment = Alignment(horizontal='center', vertical='center')
            cell_mid.border = thin_border
            
            for i, col_name in enumerate(sub_group['cols']):
                cell_bottom = ws.cell(row=3, column=col_pos + i)
                cell_bottom.value = col_name
                cell_bottom.fill = PatternFill(start_color=group['color_bottom'], 
                                              end_color=group['color_bottom'], fill_type="solid")
                cell_bottom.font = header_font_dark
                cell_bottom.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_bottom.border = thin_border
            
            col_pos += sub_span
    
    # 컬럼 너비
    column_widths = {
        '상태': 8, '신뢰도': 10, '로켓': 12, '아이허브': 12, '품번': 13, 'UPC': 15,
        '판매량': 9, '재고': 9, '재고소진일': 11, '판매가': 11, '로켓가격': 11,
        '로켓순위': 10, '아이허브순위': 13, '매출': 11, '아이템위너비율': 15,
        '판매량↔': 10, '재고↔': 9, '판매가↔': 10, '로켓가격↔': 11,
        '로켓순위↔': 11, '아이허브순위↔': 14,
        'Product_ID': 14, '로켓_Vendor': 17, '로켓_Item': 15, 
        '아이허브_Vendor': 17, '아이허브_Item': 15
    }
    
    PRODUCT_NAME_WIDTH = 60.0
    DEFAULT_WIDTH = 12
    
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        mid_header = ws.cell(row=2, column=col_idx).value
        bottom_header = ws.cell(row=3, column=col_idx).value
        
        if mid_header == '제품명':
            width = PRODUCT_NAME_WIDTH
        elif mid_header == '링크':
            width = 10
        elif bottom_header in column_widths:
            width = column_widths[bottom_header]
        else:
            width = DEFAULT_WIDTH
        
        ws.column_dimensions[col_letter].width = width
    
    # 데이터 셀 스타일
    data_start = 4
    for row_idx in range(data_start, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    # 변화량 조건부 서식
    header_names = [ws.cell(row=3, column=i).value for i in range(1, ws.max_column + 1)]
    
    def col_idx_of(name):
        try:
            return header_names.index(name) + 1
        except ValueError:
            return None
    
    # 판매량 변화
    sales_change_col = col_idx_of('판매량↔')
    if sales_change_col:
        for row_idx in range(data_start, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=sales_change_col)
            try:
                val = float(cell.value) if cell.value else 0
                if val > 0:
                    cell.fill = PatternFill(start_color=HIGHLIGHT_GREEN, 
                                          end_color=HIGHLIGHT_GREEN, fill_type="solid")
                elif val < 0:
                    cell.fill = PatternFill(start_color=HIGHLIGHT_RED, 
                                          end_color=HIGHLIGHT_RED, fill_type="solid")
            except:
                pass
    
    # 순위 변화 (양수=상승)
    for rank_col_name in ['로켓순위↔', '아이허브순위↔']:
        rank_col = col_idx_of(rank_col_name)
        if rank_col:
            for row_idx in range(data_start, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=rank_col)
                try:
                    val = float(cell.value) if cell.value else 0
                    if val > 0:
                        cell.fill = PatternFill(start_color=HIGHLIGHT_GREEN, 
                                              end_color=HIGHLIGHT_GREEN, fill_type="solid")
                    elif val < 0:
                        cell.fill = PatternFill(start_color=HIGHLIGHT_RED, 
                                              end_color=HIGHLIGHT_RED, fill_type="solid")
                except:
                    pass
    
    # 재고소진일 경고
    stockout_col = col_idx_of('재고소진일')
    if stockout_col:
        for row_idx in range(data_start, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=stockout_col)
            try:
                val = float(cell.value) if cell.value else 0
                if 0 < val < 30:
                    cell.fill = PatternFill(start_color=HIGHLIGHT_RED, 
                                          end_color=HIGHLIGHT_RED, fill_type="solid")
                elif val > 90:
                    cell.fill = PatternFill(start_color=HIGHLIGHT_YELLOW, 
                                          end_color=HIGHLIGHT_YELLOW, fill_type="solid")
            except:
                pass
    
    # 하이퍼링크
    link_columns = []
    for col_idx in range(1, ws.max_column + 1):
        mid_header = ws.cell(row=2, column=col_idx).value
        if mid_header == '링크':
            link_columns.append(col_idx)
    
    for col_idx in link_columns:
        for row_idx in range(data_start, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            url = cell.value
            if url and str(url).strip() and str(url).startswith('http'):
                cell.value = "Link"
                cell.hyperlink = str(url)
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Freeze panes
    ws.freeze_panes = ws.cell(row=4, column=9)
    
    # 자동 필터
    ws.auto_filter.ref = (
        f"{get_column_letter(1)}{3}:"
        f"{get_column_letter(ws.max_column)}{ws.max_row}"
    )
    
    wb.save(output_path)


def main():
    """메인"""
    
    # Config import
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from config.settings import Config
    
    DB_PATH = Config.INTEGRATED_DB_PATH
    OUTPUT_DIR = Config.OUTPUT_DIR
    
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    # 추이 데이터 생성
    df_trend = get_trend_data(DB_PATH)
    
    if not df_trend.empty:
        output_file = OUTPUT_DIR / f"product_dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
        create_dashboard_excel(df_trend, str(output_file))
        
        print(f"{'='*80}")
        print(f"✅ 대시보드 생성 완료")
        print(f"{'='*80}\n")
    else:
        print("\n❌ 생성할 데이터가 없습니다.\n")


if __name__ == "__main__":
    main()