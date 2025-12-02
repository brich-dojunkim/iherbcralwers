#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Price Comparison Report (Metrics 기반)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
단일 스냅샷 기준 로켓 vs 아이허브 가격 비교 리포트
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from datetime import datetime
import sys
from pathlib import Path

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from src.data_manager import DataManager
from src.metrics.core import MetricsManager


def get_data(db_path: str, target_date: str = None):
    """데이터 로드 (단일 스냅샷)"""
    
    manager = DataManager(db_path)
    metrics = MetricsManager(manager)
    
    # 단일 스냅샷 뷰
    df = metrics.get_view(
        metric_groups=['all'],  # 모든 메트릭
        n_latest=1,
        include_unmatched=True,
        compute_deltas=False
    )
    
    if df.empty:
        print("⚠️ 데이터가 없습니다.")
        return df
    
    # 시간축 suffix 제거 (단일 스냅샷이므로)
    df.columns = [col.split('__')[0] if '__' in col else col for col in df.columns]
    
    # 매출/판매량 비중 계산
    def calculate_share(colname, outname):
        total = pd.to_numeric(df[colname], errors='coerce').fillna(0).sum()
        if total <= 0:
            df[outname] = np.nan
        else:
            df[outname] = (pd.to_numeric(df[colname], errors='coerce').fillna(0) / total * 100).round(0).astype('Int64')
    
    if 'iherb_revenue' in df.columns:
        calculate_share('iherb_revenue', '매출비중')
    if 'iherb_sales_quantity' in df.columns:
        calculate_share('iherb_sales_quantity', '판매량비중')
    
    # 정수 변환
    if 'iherb_item_winner_ratio' in df.columns:
        df['iherb_item_winner_ratio'] = pd.to_numeric(df['iherb_item_winner_ratio'], errors='coerce').fillna(0).round(0).astype('Int64')
    
    print(f"\n✅ 총 {len(df):,}개 상품")
    print(f"   - 로켓 매칭: {(df['matching_status'] == '로켓매칭').sum():,}개")
    print(f"   - 미매칭 우수: {(df['matching_status'] == '미매칭').sum():,}개")
    
    # 신뢰도 분포
    matched_df = df[df['matching_status'] == '로켓매칭']
    if len(matched_df) > 0:
        conf_counts = matched_df['matching_confidence'].value_counts()
        print(f"\n   📊 매칭 신뢰도 분포:")
        for conf, count in conf_counts.items():
            pct = count / len(matched_df) * 100
            print(f"      • {conf}: {count:,}개 ({pct:.1f}%)")
    
    return df


def create_excel_report(df: pd.DataFrame, output_path: str):
    """Excel 리포트 생성"""
    
    if df.empty:
        print("❌ 데이터가 없어 엑셀 생성을 건너뜁니다.")
        return
    
    print(f"\n{'='*80}")
    print(f"📊 Excel 리포트 생성")
    print(f"{'='*80}")
    
    # 컬럼 재구성 (41개)
    output_df = pd.DataFrame({
        # 1️⃣ 기본 정보 (8개)
        '매칭상태': df.get('matching_status', np.nan),
        '신뢰도': df.get('matching_confidence', np.nan),
        '로켓_카테고리': df.get('rocket_category', np.nan),
        '아이허브_카테고리': df.get('iherb_category', np.nan),
        '로켓_링크': df.get('rocket_url', np.nan),
        '아이허브_링크': df.get('iherb_url', np.nan),
        '품번': df.get('iherb_part_number', np.nan),
        'UPC': pd.to_numeric(df.get('iherb_upc', np.nan), errors='coerce').astype('Int64'),
        
        # 2️⃣ 핵심 지표 (9개)
        '순위': df.get('rocket_rank', np.nan),
        '판매량': df.get('iherb_sales_quantity', np.nan),
        '매출(원)': df.get('iherb_revenue', np.nan),
        '아이템위너비율': df.get('iherb_item_winner_ratio', np.nan),
        '가격격차(원)': df.get('price_diff', np.nan),
        '손익분기할인율': df.get('breakeven_discount_rate', np.nan),
        '추천할인율': df.get('recommended_discount_rate', np.nan),
        '요청할인율': df.get('requested_discount_rate', np.nan),
        '유리한곳': df.get('cheaper_source', np.nan),
        
        # 3️⃣ 제품 정보 (7개)
        '로켓_제품명': df.get('rocket_product_name', np.nan),
        '아이허브_제품명': df.get('iherb_product_name', np.nan),
        'Product_ID': df.get('rocket_product_id', np.nan),
        '로켓_Vendor': df.get('rocket_vendor_id', np.nan),
        '로켓_Item': df.get('rocket_item_id', np.nan),
        '아이허브_Vendor': df.get('iherb_vendor_id', np.nan),
        '아이허브_Item': df.get('iherb_item_id', np.nan),
        
        # 4️⃣ 가격 정보 (8개)
        '정가': df.get('rocket_original_price', np.nan),
        '할인율': df.get('rocket_discount_rate', np.nan),
        '로켓가격': df.get('rocket_price', np.nan),
        '판매가': df.get('iherb_price', np.nan),
        '정가_아이허브': df.get('iherb_original_price', np.nan),
        '쿠팡추천가': df.get('iherb_recommended_price', np.nan),
        '재고': df.get('iherb_stock', np.nan),
        '판매상태': df.get('iherb_stock_status', np.nan),
        
        # 5️⃣ 판매 성과 (7개)
        '평점': df.get('rocket_rating', np.nan),
        '리뷰수': df.get('rocket_reviews', np.nan),
        '매출비중': df.get('매출비중', np.nan),
        '주문': np.nan,
        '판매량비중': df.get('판매량비중', np.nan),
        '구매전환율': np.nan,
        '취소율': np.nan,
    })
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='Price_Comparison', index=False, header=False)
    
    apply_excel_styles(output_path)
    
    print(f"\n✅ Excel 리포트 생성 완료: {output_path}")


def apply_excel_styles(output_path: str):
    """Excel 스타일 적용 (3단 헤더)"""
    
    wb = load_workbook(output_path)
    ws = list(wb.worksheets)[0]
    
    # 색상 팔레트
    INFO_DARK = "0F172A"
    INFO_MID = "475569"
    INFO_LIGHT = "E2E8F0"
    PRIMARY_DARK = "5E2A8A"
    PRIMARY_MID = "7A3EB1"
    PRIMARY_LIGHT = "D2B7E5"
    SECONDARY_DARK = "305496"
    SECONDARY_MID = "4472C4"
    SECONDARY_LIGHT = "B4C7E7"
    TERTIARY_DARK = "C55A11"
    TERTIARY_MID = "F4B084"
    TERTIARY_LIGHT = "FBE5D6"
    SUCCESS_DARK = "375623"
    SUCCESS_MID = "548235"
    SUCCESS_LIGHT = "A8D08D"
    HIGHLIGHT_GREEN = "C6EFCE"
    HIGHLIGHT_RED = "FFC7CE"
    HIGHLIGHT_YELLOW = "FFEB9C"
    
    header_font_white = Font(color="FFFFFF", bold=True, size=11)
    header_font_dark = Font(color="000000", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 컬럼 그룹 정의
    column_groups = [
        {
            'name': '기본 정보',
            'color_top': INFO_DARK,
            'color_mid': INFO_MID,
            'color_bottom': INFO_LIGHT,
            'sub_groups': [
                {'name': '매칭', 'cols': ['상태', '신뢰도']},
                {'name': '카테고리', 'cols': ['로켓', '아이허브']},
                {'name': '링크', 'cols': ['로켓', '아이허브']},
                {'name': '상품번호', 'cols': ['품번', 'UPC']}
            ]
        },
        {
            'name': '핵심 지표',
            'color_top': PRIMARY_DARK,
            'color_mid': PRIMARY_MID,
            'color_bottom': PRIMARY_LIGHT,
            'sub_groups': [
                {'name': '로켓', 'cols': ['순위']},
                {'name': '아이허브', 'cols': ['판매량', '매출(원)', '아이템위너비율']},
                {'name': '종합', 'cols': ['가격격차(원)', '손익분기할인율', '추천할인율', '요청할인율', '유리한곳']}
            ]
        },
        {
            'name': '제품 정보',
            'color_top': SECONDARY_DARK,
            'color_mid': SECONDARY_MID,
            'color_bottom': SECONDARY_LIGHT,
            'sub_groups': [
                {'name': '제품명', 'cols': ['로켓', '아이허브']},
                {'name': '상품 ID', 'cols': ['Product_ID', '로켓_Vendor', '로켓_Item', '아이허브_Vendor', '아이허브_Item']}
            ]
        },
        {
            'name': '가격 정보',
            'color_top': TERTIARY_DARK,
            'color_mid': TERTIARY_MID,
            'color_bottom': TERTIARY_LIGHT,
            'sub_groups': [
                {'name': '로켓직구', 'cols': ['정가', '할인율', '로켓가격']},
                {'name': '아이허브', 'cols': ['판매가', '정가', '쿠팡추천가', '재고', '판매상태']}
            ]
        },
        {
            'name': '판매 성과',
            'color_top': SUCCESS_DARK,
            'color_mid': SUCCESS_MID,
            'color_bottom': SUCCESS_LIGHT,
            'sub_groups': [
                {'name': '로켓', 'cols': ['평점', '리뷰수']},
                {'name': '아이허브', 'cols': ['매출비중', '주문', '판매량비중', '구매전환율', '취소율']}
            ]
        }
    ]
    
    # 3개 행 삽입
    ws.insert_rows(1, 3)
    
    # 헤더 작성
    col_pos = 1
    
    for group in column_groups:
        group_name = group['name']
        color_top = group['color_top']
        color_mid = group['color_mid']
        color_bottom = group['color_bottom']
        sub_groups = group['sub_groups']
        
        total_span = sum(len(sg['cols']) for sg in sub_groups)
        
        # 상위 헤더
        ws.merge_cells(start_row=1, start_column=col_pos, end_row=1, end_column=col_pos + total_span - 1)
        cell_top = ws.cell(row=1, column=col_pos)
        cell_top.value = group_name
        cell_top.fill = PatternFill(start_color=color_top, end_color=color_top, fill_type="solid")
        cell_top.font = header_font_white
        cell_top.alignment = Alignment(horizontal='center', vertical='center')
        cell_top.border = thin_border
        
        # 중간 헤더
        for sub_group in sub_groups:
            sub_name = sub_group['name']
            sub_cols = sub_group['cols']
            sub_span = len(sub_cols)
            
            ws.merge_cells(start_row=2, start_column=col_pos, end_row=2, end_column=col_pos + sub_span - 1)
            cell_mid = ws.cell(row=2, column=col_pos)
            cell_mid.value = sub_name
            cell_mid.fill = PatternFill(start_color=color_mid, end_color=color_mid, fill_type="solid")
            cell_mid.font = header_font_white
            cell_mid.alignment = Alignment(horizontal='center', vertical='center')
            cell_mid.border = thin_border
            
            # 하위 헤더
            for i, col_name in enumerate(sub_cols):
                cell_bottom = ws.cell(row=3, column=col_pos + i)
                cell_bottom.value = col_name
                cell_bottom.fill = PatternFill(start_color=color_bottom, end_color=color_bottom, fill_type="solid")
                cell_bottom.font = header_font_dark
                cell_bottom.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_bottom.border = thin_border
            
            col_pos += sub_span
    
    # 컬럼 인덱스 찾기
    header_names = [cell.value for cell in ws[3]]
    
    def col_idx_of(name):
        try:
            return header_names.index(name) + 1
        except ValueError:
            return None
    
    # 컬럼 너비 설정
    column_widths = {
        '상태': 7.9,
        '신뢰도': 9.71,
        '로켓': 12.86,
        '아이허브': 11.00,
        '품번': 12.71,
        'UPC': 15.00,
        '순위': 7.86,
        '판매량': 9.00,
        '매출(원)': 10.14,
        '아이템위너비율': 15.57,
        '가격격차(원)': 13.43,
        '손익분기할인율': 15.57,
        '추천할인율': 12.29,
        '요청할인율': 12.29,
        '유리한곳': 10.57,
        'Product_ID': 14.14,
        '로켓_Vendor': 17.29,
        '로켓_Item': 15.14,
        '아이허브_Vendor': 17.29,
        '아이허브_Item': 15.14,
        '정가': 8.86,
        '할인율': 9.00,
        '로켓가격': 10.57,
        '판매가': 10.57,
        '쿠팡추천가': 12.29,
        '재고': 7.29,
        '판매상태': 10.57,
        '평점': 8.86,
        '리뷰수': 9.00,
        '매출비중': 10.57,
        '주문': 7.29,
        '판매량비중': 12.29,
        '구매전환율': 13.00,
        '취소율': 9.00,
    }
    
    DEFAULT_WIDTH = 12
    PRODUCT_NAME_WIDTH = 60.0
    
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        
        mid_header = ws.cell(row=2, column=col_idx).value
        if mid_header is None and col_idx > 1:
            mid_header = ws.cell(row=2, column=col_idx - 1).value
        
        bottom_header = ws.cell(row=3, column=col_idx).value
        
        if mid_header == '제품명':
            width = PRODUCT_NAME_WIDTH
        elif mid_header == '링크' and bottom_header == '로켓':
            width = 7.29
        elif mid_header == '링크' and bottom_header == '아이허브':
            width = 10.57
        elif bottom_header in column_widths:
            width = column_widths[bottom_header]
        else:
            width = DEFAULT_WIDTH
        
        ws.column_dimensions[col_letter].width = width
    
    # 데이터 셀 기본 스타일
    data_actual_start = 4
    max_col = ws.max_column
    
    for row_idx in range(data_actual_start, ws.max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
    
    # 조건부 서식: 아이템위너비율 >= 30%
    winner_col = col_idx_of('아이템위너비율')
    if winner_col:
        for row_idx in range(data_actual_start, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=winner_col)
            try:
                val = float(cell.value) if cell.value else 0
                if val >= 30:
                    cell.fill = PatternFill(start_color=HIGHLIGHT_GREEN, end_color=HIGHLIGHT_GREEN, fill_type="solid")
            except:
                pass
    
    # 조건부 서식: 손익분기할인율 > 0
    breakeven_col = col_idx_of('손익분기할인율')
    if breakeven_col:
        for row_idx in range(data_actual_start, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=breakeven_col)
            try:
                val = float(cell.value) if cell.value else 0
                if val > 0:
                    cell.fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
            except:
                pass
    
    # 조건부 서식: 추천할인율 > 0
    recommended_col = col_idx_of('추천할인율')
    if recommended_col:
        for row_idx in range(data_actual_start, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=recommended_col)
            try:
                val = float(cell.value) if cell.value else 0
                if val > 0:
                    cell.fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
            except:
                pass
    
    # 조건부 서식: 요청할인율 > 0
    requested_col = col_idx_of('요청할인율')
    if requested_col:
        for row_idx in range(data_actual_start, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=requested_col)
            try:
                val = float(cell.value) if cell.value else 0
                if val > 0:
                    cell.fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
            except:
                pass
    
    # 조건부 서식: 가격격차 + 유리한곳 연동
    price_diff_col = col_idx_of('가격격차(원)')
    cheaper_col = col_idx_of('유리한곳')
    if price_diff_col and cheaper_col:
        for row_idx in range(data_actual_start, ws.max_row + 1):
            cheaper_value = ws.cell(row=row_idx, column=cheaper_col).value
            if cheaper_value == '아이허브':
                ws.cell(row=row_idx, column=price_diff_col).fill = PatternFill(start_color=HIGHLIGHT_GREEN, end_color=HIGHLIGHT_GREEN, fill_type="solid")
            elif cheaper_value == '로켓직구':
                ws.cell(row=row_idx, column=price_diff_col).fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
    
    # 조건부 서식: 신뢰도
    conf_col = col_idx_of('신뢰도')
    if conf_col:
        for row_idx in range(data_actual_start, ws.max_row + 1):
            conf_value = ws.cell(row=row_idx, column=conf_col).value
            cell = ws.cell(row=row_idx, column=conf_col)
            if conf_value == 'High':
                cell.fill = PatternFill(start_color=HIGHLIGHT_GREEN, end_color=HIGHLIGHT_GREEN, fill_type="solid")
            elif conf_value == 'Medium':
                cell.fill = PatternFill(start_color=HIGHLIGHT_YELLOW, end_color=HIGHLIGHT_YELLOW, fill_type="solid")
            elif conf_value == 'Low':
                cell.fill = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
    
    # 하이퍼링크
    link_columns = []
    for col_idx in range(1, max_col + 1):
        mid_header = ws.cell(row=2, column=col_idx).value
        
        if mid_header == '링크':
            link_columns.append(col_idx)
        elif mid_header is None and col_idx > 1:
            left_mid_header = ws.cell(row=2, column=col_idx - 1).value
            if left_mid_header == '링크':
                link_columns.append(col_idx)
    
    for col_idx in link_columns:
        for row_idx in range(data_actual_start, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            url = cell.value
            if url and str(url).strip() and str(url).startswith('http'):
                cell.value = "Link"
                cell.hyperlink = str(url)
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Freeze Panes
    freeze_col = 17
    ws.freeze_panes = ws.cell(row=4, column=freeze_col)
    
    # 데이터바: 매출비중, 판매량비중
    share_cols = ['매출비중', '판매량비중']
    for share_col_name in share_cols:
        share_col_idx = col_idx_of(share_col_name)
        if share_col_idx:
            col_letter = get_column_letter(share_col_idx)
            rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='num', end_value=100,
                color="63C384"
            )
            ws.conditional_formatting.add(
                f'{col_letter}{data_actual_start}:{col_letter}{ws.max_row}',
                rule
            )
    
    # 자동 필터
    ws.auto_filter.ref = (
        f"{get_column_letter(1)}{3}:"
        f"{get_column_letter(ws.max_column)}{ws.max_row}"
    )
    
    wb.save(output_path)


def main():
    from config.settings import Config
    
    DB_PATH = Config.INTEGRATED_DB_PATH
    OUTPUT_DIR = Config.OUTPUT_DIR
    
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    df = get_data(DB_PATH)
    
    if not df.empty:
        output_file = OUTPUT_DIR / f"price_comparison_{datetime.now().strftime('%Y%m%d')}.xlsx"
        create_excel_report(df, str(output_file))
    else:
        print("\n❌ 생성할 데이터가 없습니다.")


if __name__ == "__main__":
    main()