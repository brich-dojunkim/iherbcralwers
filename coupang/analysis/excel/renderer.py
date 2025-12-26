#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel Renderer
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Excel 렌더링 엔진 (컬럼-값 매핑 책임)

🔥 수정사항:
  - 3단 헤더 → 2단 헤더로 변경
  - 1단: 그룹명
  - 2단: 컬럼명 (서브그룹 제거)
  - 모든 row 번호 -1 조정
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

from .types import ExcelConfig
from .constants import COLOR_SCHEMES


class ExcelRenderer:
    """Excel 렌더러"""
    
    def __init__(self, output_path: str, sheet_name: str = 'Sheet1'):
        self.output_path = Path(output_path)
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
    
    def render(self, df: pd.DataFrame, config: ExcelConfig) -> dict:
        """DataFrame → Excel 렌더링
        
        🔥 Excel Layer 책임: config.columns 순서대로 컬럼-값 자동 매핑
        
        Returns:
            {'success': bool, 'path': str, 'rows': int, 'cols': int, 'error': str}
        """
        try:
            # 🔥 Step 0: config 순서대로 컬럼 자동 정렬
            ordered_columns = [col.name for col in config.columns]
            
            # df에 없는 컬럼 체크
            missing = [c for c in ordered_columns if c not in df.columns]
            if missing:
                print(f"⚠️  경고: DataFrame에 없는 컬럼: {missing}")
                for col in missing:
                    df[col] = pd.NA
            
            # config.columns 순서대로 재정렬 (Analysis 순서 무관)
            df = df[ordered_columns].copy()
            
            print(f"[RENDERER] 0/9 컬럼 정렬 완료 (config 순서 보장)")
            
            print(f"[RENDERER] 1/9 데이터 쓰기... ({len(df):,}행)")
            # 1. 데이터 쓰기
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=self.sheet_name, index=False, header=False)
            print(f"[RENDERER] 1/9 완료")
            
            print(f"[RENDERER] 2/9 파일 로드...")
            # 2. 파일 로드 - 🔥 2행 삽입 (2단 헤더)
            self.wb = load_workbook(self.output_path)
            self.ws = self.wb[self.sheet_name]
            self.ws.insert_rows(1, 2)  # 🔥 수정: 3 → 2
            print(f"[RENDERER] 2/9 완료")
            
            print(f"[RENDERER] 3/9 헤더 렌더링...")
            # 3. 헤더 - 🔥 2단 헤더
            self._render_headers(config.groups)
            print(f"[RENDERER] 3/9 완료")
            
            print(f"[RENDERER] 4/9 컬럼 너비...")
            # 4. 컬럼 너비
            self._set_column_widths(config.columns)
            print(f"[RENDERER] 4/9 완료")
            
            print(f"[RENDERER] 5/9 데이터 영역 스타일...")
            # 5. 데이터 영역
            self._style_data_area(config.columns)
            print(f"[RENDERER] 5/9 완료")
            
            print(f"[RENDERER] 6/9 조건부 서식... ({len(config.conditional_rules)}개 규칙)")
            # 6. 조건부 서식
            if config.conditional_rules:
                self._apply_conditional_rules(config.conditional_rules)
            print(f"[RENDERER] 6/9 완료")
            
            print(f"[RENDERER] 7/9 데이터바...")
            # 7. 🆕 데이터바
            self._apply_data_bars()
            print(f"[RENDERER] 7/9 완료")
            
            print(f"[RENDERER] 8/9 링크 처리...")
            # 8. 링크
            self._apply_links()
            print(f"[RENDERER] 8/9 완료")
            
            print(f"[RENDERER] 9/9 UI 설정...")
            # 9. UI
            if config.freeze_panes:
                row, col = config.freeze_panes
                self.ws.freeze_panes = self.ws.cell(row, col)  # 🔥 그대로 사용 (price_comparison_2에서 조정됨)
            
            if config.auto_filter:
                self.ws.auto_filter.ref = (
                    f"A2:{get_column_letter(self.ws.max_column)}{self.ws.max_row}"  # 🔥 A3 → A2
                )
            
            # 저장
            self.wb.save(self.output_path)
            print(f"[RENDERER] 9/9 완료")
            
            return {
                'success': True,
                'path': str(self.output_path),
                'rows': len(df),
                'cols': len(df.columns),
                'error': None
            }
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'path': str(self.output_path),
                'rows': 0,
                'cols': 0,
                'error': str(e)
            }
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Private Methods
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def _render_headers(self, groups):
        """🔥 2단 헤더 렌더링 (서브그룹 제거)"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        col_pos = 1
        
        for group in groups:
            colors = COLOR_SCHEMES[group.color_scheme]
            
            # 전체 컬럼 수 계산 (서브그룹의 모든 컬럼 합치기)
            all_columns = []
            for sg in group.sub_groups:
                all_columns.extend(sg.columns)
            total_span = len(all_columns)
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 1단: 그룹 헤더
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            for i in range(col_pos, col_pos + total_span):
                cell = self.ws.cell(1, i)
                cell.fill = PatternFill(start_color=colors["top"], end_color=colors["top"], fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            if total_span > 1:
                self.ws.merge_cells(start_row=1, start_column=col_pos, end_row=1, end_column=col_pos + total_span - 1)
            
            self.ws.cell(1, col_pos).value = group.name
            
            # 양 끝 강조
            self.ws.cell(1, col_pos).border = Border(
                left=Side(style='medium'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            self.ws.cell(1, col_pos + total_span - 1).border = Border(
                left=Side(style='thin'), right=Side(style='medium'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 2단: 컬럼명 (서브그룹 건너뛰고 바로 컬럼)
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            for i, col_name in enumerate(all_columns):
                cell = self.ws.cell(2, col_pos + i)  # 🔥 3 → 2
                cell.value = col_name
                cell.fill = PatternFill(start_color=colors["bottom"], end_color=colors["bottom"], fill_type="solid")
                cell.font = Font(color="000000", bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            col_pos += total_span
    
    def _set_column_widths(self, columns):
        """컬럼 너비 설정"""
        for col_idx, col_spec in enumerate(columns, 1):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = col_spec.width
    
    def _style_data_area(self, columns):
        """데이터 영역 스타일 (테두리, 서식, 정렬)"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 🔥 수정: 4 → 3
        for row_idx in range(3, self.ws.max_row + 1):
            for col_idx, col_spec in enumerate(columns, 1):
                cell = self.ws.cell(row_idx, col_idx)
                cell.border = border
                cell.number_format = col_spec.number_format
                cell.alignment = Alignment(
                    horizontal=col_spec.alignment,
                    vertical='center',
                    wrap_text=False
                )
    
    def _apply_conditional_rules(self, rules):
        """조건부 서식 적용 (최적화: 컬럼별 한 번만 순회)"""
        # 컬럼별로 규칙 그룹핑
        rules_by_column = {}
        for rule in rules:
            if rule.column not in rules_by_column:
                rules_by_column[rule.column] = []
            rules_by_column[rule.column].append(rule)
        
        # 컬럼별 처리
        for col_name, col_rules in rules_by_column.items():
            col_idx = self._find_column(col_name)
            if not col_idx:
                continue
            
            # 🔥 수정: 4 → 3
            # 해당 컬럼만 한 번 순회
            for row_idx in range(3, self.ws.max_row + 1):
                cell = self.ws.cell(row_idx, col_idx)
                value = cell.value
                
                # 모든 규칙 체크 (첫 번째 매칭되는 것 적용)
                for rule in col_rules:
                    try:
                        if rule.condition(value):
                            if rule.fill_color:
                                cell.fill = PatternFill(
                                    start_color=rule.fill_color,
                                    end_color=rule.fill_color,
                                    fill_type="solid"
                                )
                            if rule.font_color:
                                cell.font = Font(color=rule.font_color)
                            break  # 첫 매칭만
                    except:
                        pass
    
    def _apply_data_bars(self):
        """🆕 데이터바 자동 적용 (비중 컬럼 감지)"""
        # 🔥 수정: 3행 → 2행 (컬럼명 위치)
        # 2행에서 '비중' 포함 컬럼 찾기
        for col_idx in range(1, self.ws.max_column + 1):
            col_name = self.ws.cell(2, col_idx).value
            
            if col_name and '비중' in str(col_name):
                col_letter = get_column_letter(col_idx)
                rule = DataBarRule(
                    start_type='num', start_value=0,
                    end_type='num', end_value=100,
                    color="63C384"
                )
                # 🔥 수정: 4 → 3
                self.ws.conditional_formatting.add(
                    f'{col_letter}3:{col_letter}{self.ws.max_row}',
                    rule
                )
    
    def _apply_links(self):
        """하이퍼링크 처리"""
        # 🔥 수정: 3행 → 2행 (컬럼명 위치)
        # 2행에서 링크 컬럼 찾기
        link_columns = []
        for col_idx in range(1, self.ws.max_column + 1):
            col_name = self.ws.cell(2, col_idx).value
            if col_name and ('링크' in str(col_name) or 'url' in str(col_name).lower()):
                link_columns.append(col_idx)
        
        # 🔥 수정: 4 → 3
        # 링크 처리
        for col_idx in link_columns:
            for row_idx in range(3, self.ws.max_row + 1):
                cell = self.ws.cell(row_idx, col_idx)
                url = cell.value
                
                if url and str(url).startswith('http'):
                    cell.hyperlink = str(url)
                    cell.value = "Link"
                    cell.font = Font(color="0563C1", underline="single")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _find_column(self, col_name: str) -> int:
        """컬럼명으로 인덱스 찾기 (2행 기준)"""
        # 🔥 수정: 3행 → 2행
        for col_idx in range(1, self.ws.max_column + 1):
            if self.ws.cell(2, col_idx).value == col_name:
                return col_idx
        return None