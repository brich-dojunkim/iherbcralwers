#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import sqlite3
import sys
from pathlib import Path

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from src.data_manager import DataManager
from src.metrics.schema import ACTION_METRICS  # 🔹 액션 메트릭 정의 가져오기


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 액션 메트릭 → 엑셀 컬럼 라벨 매핑
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ACTION_LABELS = {
    "requested_discount_rate": "요청할인율",
    "recommended_discount_rate": "추천할인율",
    "breakeven_discount_rate": "손익분기할인율",
    "cheaper_source": "유리한곳",
    "price_diff": "가격격차",
    # 필요하면 이후에 price_diff_pct 등도 붙일 수 있음
}


def get_data(db_path: str):
    """데이터 로드 + 전일/오늘 비교용 컬럼 생성 + 스냅샷 날짜 반환"""
    conn = sqlite3.connect(db_path)

    # 최신 2개 스냅샷 (id + date)
    cursor = conn.execute(
        "SELECT id, snapshot_date FROM snapshots ORDER BY id DESC LIMIT 2"
    )
    rows = cursor.fetchall()
    if len(rows) < 2:
        conn.close()
        return pd.DataFrame(), None, None

    current_id, current_date = rows[0]
    prev_id, prev_date = rows[1]

    manager = DataManager(db_path)
    df_curr = manager.get_integrated_df(snapshot_id=current_id, include_unmatched=True)
    df_prev = manager.get_integrated_df(snapshot_id=prev_id, include_unmatched=True)

    # 미매칭 로켓 제외 (로켓은 있는데 아이허브 없는 케이스)
    if not df_curr.empty:
        df_curr = df_curr[
            ~(
                df_curr["rocket_vendor_id"].notna()
                & df_curr["iherb_vendor_id"].isna()
            )
        ]

    if df_curr.empty:
        conn.close()
        return pd.DataFrame(), current_date, prev_date

    # 전일 데이터(이전 스냅샷)에서 판매량/가격/위너비율만 추출
    needed_cols = [
        "iherb_vendor_id",
        "iherb_sales_quantity",
        "iherb_price",
        "iherb_item_winner_ratio",
    ]
    existing_prev_cols = [c for c in needed_cols if c in df_prev.columns]

    if not existing_prev_cols:
        # 이전 스냅샷에 필요한 컬럼이 하나도 없으면, 전일 비교 없이 진행
        df = df_curr.copy()
        df["discount_prev"] = np.nan
        df["discount_curr"] = np.nan
        conn.close()
        return df, current_date, prev_date

    prev = df_prev[existing_prev_cols].copy()

    # 안전하게 컬럼명 보정
    rename_map = {}
    if "iherb_vendor_id" in prev.columns:
        rename_map["iherb_vendor_id"] = "iherb_vendor_id"
    if "iherb_sales_quantity" in prev.columns:
        rename_map["iherb_sales_quantity"] = "sales_prev"
    if "iherb_price" in prev.columns:
        rename_map["iherb_price"] = "price_prev"
    if "iherb_item_winner_ratio" in prev.columns:
        rename_map["iherb_item_winner_ratio"] = "winner_prev"

    prev = prev.rename(columns=rename_map)

    # 벤더 기준 중복 제거 (마지막 값 유지)
    if "iherb_vendor_id" in prev.columns:
        prev = prev.drop_duplicates("iherb_vendor_id", keep="last")

    # 병합 (오늘 스냅샷 기준, 전일 값 붙이기)
    if "iherb_vendor_id" in df_curr.columns:
        df = df_curr.merge(prev, on="iherb_vendor_id", how="left")
    else:
        # vendor_id 자체가 없으면 그냥 현재 데이터만 사용
        df = df_curr.copy()
        df["sales_prev"] = np.nan
        df["price_prev"] = np.nan
        df["winner_prev"] = np.nan

    # 할인율 계산 함수
    def calc_discount(orig, sale):
        o = pd.to_numeric(orig, errors="coerce")
        s = pd.to_numeric(sale, errors="coerce")
        result = pd.Series(np.nan, index=o.index)
        mask = (o > 0) & (s > 0)
        result[mask] = ((o[mask] - s[mask]) / o[mask] * 100).round(1)
        return result

    # 할인율(전일/오늘) – D-1 / D 기준
    df["discount_prev"] = calc_discount(df["iherb_original_price"], df["price_prev"])
    df["discount_curr"] = calc_discount(df["iherb_original_price"], df["iherb_price"])

    conn.close()
    return df, current_date, prev_date


def create_excel(df: pd.DataFrame, output: str, curr_date: str, prev_date: str):
    """Price Agent Excel 생성"""
    if df.empty:
        return

    # 문자열 → date 객체
    curr_d = datetime.strptime(curr_date, "%Y-%m-%d").date()
    prev_d = datetime.strptime(prev_date, "%Y-%m-%d").date()

    # 🔹 할인율/가격: D / D-1
    col_판매가_오늘 = f"판매가\n({curr_d})"
    col_판매가_전일 = f"판매가\n({prev_d})"
    col_할인율_오늘 = f"할인율\n({curr_d})"
    col_할인율_전일 = f"할인율\n({prev_d})"

    # 🔹 판매량/위너: D-1 / D-2
    날짜_판매량_어제 = curr_d - timedelta(days=1)
    날짜_판매량_그제 = prev_d - timedelta(days=1)
    col_판매량_어제 = f"판매량\n({날짜_판매량_어제})"
    col_판매량_그제 = f"판매량\n({날짜_판매량_그제})"
    col_위너비율_어제 = f"위너비율\n({날짜_판매량_어제})"
    col_위너비율_그제 = f"위너비율\n({날짜_판매량_그제})"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ① 기본 식별자 + 액션 메트릭 (metrics.schema 기반)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    data = {
        # 코어
        "매칭상태": df.get("matching_status"),
        "품번": df.get("iherb_part_number"),
        "Product_ID": df.get("product_id"),
    }

    # 액션 메트릭: ACTION_METRICS 순서대로, 매핑된 라벨만 추가
    for metric_key in ACTION_METRICS:
        if metric_key in ACTION_LABELS and metric_key in df.columns:
            label = ACTION_LABELS[metric_key]
            data[label] = df[metric_key]

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ② 전일 대비 변화
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 할인율Δ (NaType 안전 처리)
    discount_curr = pd.to_numeric(df.get("discount_curr"), errors="coerce")
    discount_prev = pd.to_numeric(df.get("discount_prev"), errors="coerce")
    delta_discount = (discount_curr - discount_prev).round(1)

    # 판매량Δ
    sales_curr = pd.to_numeric(df.get("iherb_sales_quantity"), errors="coerce")
    sales_prev = pd.to_numeric(df.get("sales_prev"), errors="coerce")
    delta_sales = (sales_curr - sales_prev).round(0)

    # 위너비율Δ
    winner_curr = pd.to_numeric(df.get("iherb_item_winner_ratio"), errors="coerce")
    winner_prev = pd.to_numeric(df.get("winner_prev"), errors="coerce")
    delta_winner = (winner_curr - winner_prev).round(1)

    data["할인율Δ"] = delta_discount
    data["판매량Δ"] = delta_sales
    data["위너비율Δ"] = delta_winner

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ③ 가격/할인율 상태 (이전 → 오늘)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    data["정가"] = df.get("iherb_original_price")
    data[col_판매가_전일] = df.get("price_prev")
    data[col_판매가_오늘] = df.get("iherb_price")
    data[col_할인율_전일] = df.get("discount_prev")
    data[col_할인율_오늘] = df.get("discount_curr")
    data["로켓_판매가"] = df.get("rocket_price")

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ④ 판매/위너 스냅샷 (이전 → 오늘 기준에 가까운 날짜)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    data[col_판매량_그제] = df.get("sales_prev")
    data[col_판매량_어제] = df.get("iherb_sales_quantity")
    data[col_위너비율_그제] = df.get("winner_prev")
    data[col_위너비율_어제] = df.get("iherb_item_winner_ratio")

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ⑤ 상세 메타 정보
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    data["제품명_아이허브"] = df.get("iherb_product_name")
    data["제품명_로켓"] = df.get("rocket_product_name")
    data["카테고리_아이허브"] = df.get("iherb_category")
    data["카테고리_로켓"] = df.get("rocket_category")
    data["링크_아이허브"] = df.get("iherb_url")
    data["링크_로켓"] = df.get("rocket_url")
    data["Vendor_아이허브"] = df.get("iherb_vendor_id")
    data["Vendor_로켓"] = df.get("rocket_vendor_id")
    data["Item_아이허브"] = df.get("iherb_item_id")
    data["Item_로켓"] = df.get("rocket_item_id")
    data["순위_아이허브"] = (
        df.get("iherb_sales_quantity")
        .rank(method="min", ascending=False)
        .round(0)
    )
    data["순위_로켓"] = df.get("rocket_rank")
    data["평점"] = df.get("rocket_rating")
    data["리뷰수"] = df.get("rocket_reviews")

    out = pd.DataFrame(data)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Price_Agent", index=False, header=False)

    style(
        path=output,
        curr_d=curr_d,
        prev_d=prev_d,
        날짜_판매량_어제=날짜_판매량_어제,
        날짜_판매량_그제=날짜_판매량_그제,
    )


def style(path: str, curr_d, prev_d, 날짜_판매량_어제, 날짜_판매량_그제):
    """Excel 스타일 적용 (헤더/정렬/링크 등)"""
    wb = load_workbook(path)
    ws = wb["Price_Agent"]
    ws.insert_rows(1, 2)

    BG = "F8FAFC"
    BORDER = "E2E8F0"
    GREEN = "C6EFCE"
    RED = "FFC7CE"

    # 동적 컬럼명 (create_excel과 동일 로직, 줄바꿈 포함)
    col_판매가_오늘 = f"판매가\n({curr_d})"
    col_판매가_전일 = f"판매가\n({prev_d})"
    col_할인율_오늘 = f"할인율\n({curr_d})"
    col_할인율_전일 = f"할인율\n({prev_d})"
    col_판매량_어제 = f"판매량\n({날짜_판매량_어제})"
    col_판매량_그제 = f"판매량\n({날짜_판매량_그제})"
    col_위너비율_어제 = f"위너비율\n({날짜_판매량_어제})"
    col_위너비율_그제 = f"위너비율\n({날짜_판매량_그제})"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 액션 서브컬럼(라벨) 동적 생성
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    action_subcols = []
    for metric_key in ACTION_METRICS:
        label = ACTION_LABELS.get(metric_key)
        if label:
            action_subcols.append(("할인전략", label))

    # 컬럼 그룹 정의 (DataFrame 생성 순서와 동일해야 함)
    cols = [
        # ① 기본
        ("코어", "매칭상태"),
        ("코어", "품번"),
        ("코어", "Product_ID"),
        # ② 액션 (metrics 기반)
        *action_subcols,
        # ③ 변화
        ("변화", "할인율Δ"),
        ("변화", "판매량Δ"),
        ("변화", "위너비율Δ"),
        # ④ 가격/할인율 상태 (전일 → 오늘)
        ("가격상태", "정가"),
        ("가격상태", col_판매가_전일),
        ("가격상태", col_판매가_오늘),
        ("가격상태", col_할인율_전일),
        ("가격상태", col_할인율_오늘),
        ("가격상태", "로켓_판매가"),
        # ⑤ 판매/위너 스냅샷 (그제 → 어제)
        ("판매/위너", col_판매량_그제),
        ("판매/위너", col_판매량_어제),
        ("판매/위너", col_위너비율_그제),
        ("판매/위너", col_위너비율_어제),
        # ⑥ 메타
        ("제품명", "제품명_아이허브"),
        ("제품명", "제품명_로켓"),
        ("카테고리", "카테고리_아이허브"),
        ("카테고리", "카테고리_로켓"),
        ("링크", "링크_아이허브"),
        ("링크", "링크_로켓"),
        ("ID", "Vendor_아이허브"),
        ("ID", "Vendor_로켓"),
        ("ID", "Item_아이허브"),
        ("ID", "Item_로켓"),
        ("순위", "순위_아이허브"),
        ("순위", "순위_로켓"),
        ("평가", "평점"),
        ("평가", "리뷰수"),
    ]

    # 2행: 하위 헤더
    for i, (_, sub) in enumerate(cols, 1):
        c = ws.cell(2, i)
        c.value = sub
        c.fill = PatternFill(start_color=BG, end_color=BG, fill_type="solid")
        c.font = Font(color="1E293B", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(
            left=Side(style="thin", color=BORDER),
            right=Side(style="thin", color=BORDER),
            top=Side(style="thin", color=BORDER),
            bottom=Side(style="thin", color=BORDER),
        )

    # 1행: 그룹 헤더
    groups = {}
    for i, (grp, _) in enumerate(cols, 1):
        if grp not in groups:
            groups[grp] = [i, i]
        else:
            groups[grp][1] = i

    for grp, (start, end) in groups.items():
        if start < end:
            ws.merge_cells(
                start_row=1, start_column=start, end_row=1, end_column=end
            )

        for i in range(start, end + 1):
            c = ws.cell(1, i)
            c.fill = PatternFill(start_color=BG, end_color=BG, fill_type="solid")
            c.font = Font(color="1E293B", bold=True, size=10)
            c.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            # 그룹 양 끝 테두리 강조
            if i == start:
                c.border = Border(
                    left=Side(style="medium", color=BORDER),
                    right=Side(style="thin", color=BORDER),
                    top=Side(style="thin", color=BORDER),
                    bottom=Side(style="thin", color=BORDER),
                )
            elif i == end:
                c.border = Border(
                    left=Side(style="thin", color=BORDER),
                    right=Side(style="medium", color=BORDER),
                    top=Side(style="thin", color=BORDER),
                    bottom=Side(style="thin", color=BORDER),
                )
            else:
                c.border = Border(
                    left=Side(style="thin", color=BORDER),
                    right=Side(style="thin", color=BORDER),
                    top=Side(style="thin", color=BORDER),
                    bottom=Side(style="thin", color=BORDER),
                )

        ws.cell(1, start).value = grp

    # 컬럼 너비
    widths = {
        "매칭상태": 10,
        "품번": 15,
        "Product_ID": 15,
        "요청할인율": 14,
        "추천할인율": 14,
        "손익분기할인율": 16,
        "유리한곳": 12,
        "가격격차": 12,
        "할인율Δ": 12,
        "판매량Δ": 12,
        "위너비율Δ": 12,
        "정가": 12,
        col_판매가_오늘: 14,
        col_판매가_전일: 14,
        col_할인율_오늘: 14,
        col_할인율_전일: 14,
        "로켓_판매가": 14,
        col_판매량_어제: 14,
        col_판매량_그제: 14,
        col_위너비율_어제: 16,
        col_위너비율_그제: 16,
        "제품명_아이허브": 60,
        "제품명_로켓": 60,
        "카테고리_아이허브": 16,
        "카테고리_로켓": 16,
        "링크_아이허브": 10,
        "링크_로켓": 10,
        "Vendor_아이허브": 18,
        "Vendor_로켓": 18,
        "Item_아이허브": 16,
        "Item_로켓": 16,
        "순위_아이허브": 14,
        "순위_로켓": 14,
        "평점": 10,
        "리뷰수": 12,
    }

    for i, (_, sub) in enumerate(cols, 1):
        key = sub
        w = widths.get(key, 12)
        ws.column_dimensions[get_column_letter(i)].width = w

    # 데이터 영역 스타일
    for row in range(3, ws.max_row + 1):
        for i, (grp, sub) in enumerate(cols, 1):
            c = ws.cell(row, i)
            c.border = Border(
                left=Side(style="thin", color=BORDER),
                right=Side(style="thin", color=BORDER),
                top=Side(style="thin", color=BORDER),
                bottom=Side(style="thin", color=BORDER),
            )

            # 정렬 규칙
            if grp in ["가격상태", "할인전략", "판매/위너", "변화"]:
                c.alignment = Alignment(
                    horizontal="right", vertical="center", wrap_text=False
                )
            else:
                c.alignment = Alignment(vertical="center", wrap_text=False)

            # 숫자 서식
            if sub in [
                "매칭상태",
                "품번",
                "Product_ID",
                "유리한곳",
                "Vendor_아이허브",
                "Vendor_로켓",
                "Item_아이허브",
                "Item_로켓",
            ]:
                c.number_format = "@"
            elif sub in [
                "정가",
                col_판매가_오늘,
                col_판매가_전일,
                "로켓_판매가",
                "가격격차",
                "리뷰수",
            ]:
                c.number_format = "#,##0"
            elif sub in [col_판매량_어제, col_판매량_그제]:
                c.number_format = "#,##0"
            elif sub in [
                col_할인율_오늘,
                col_할인율_전일,
                "요청할인율",
                "추천할인율",
                "손익분기할인율",
                col_위너비율_어제,
                col_위너비율_그제,
            ]:
                # 값이 숫자인 경우에만 % 포맷 적용
                try:
                    v = float(c.value)
                    if not np.isnan(v):
                        c.number_format = '0.0"%"'
                except Exception:
                    pass
            elif sub in ["할인율Δ", "위너비율Δ"]:
                c.number_format = "0.0"
            elif sub == "판매량Δ":
                c.number_format = "#,##0"
            elif sub == "평점":
                c.number_format = "0.0"
            elif grp == "순위":
                c.number_format = "0"

            # 변화량 색상 (Δ)
            if "Δ" in sub and c.value not in (None, ""):
                try:
                    v = float(c.value)
                    if v > 0:
                        c.fill = PatternFill(
                            start_color=GREEN, end_color=GREEN, fill_type="solid"
                        )
                    elif v < 0:
                        c.fill = PatternFill(
                            start_color=RED, end_color=RED, fill_type="solid"
                        )
                except Exception:
                    pass

            # 유리한곳 색상
            if sub == "유리한곳":
                if c.value == "아이허브":
                    c.fill = PatternFill(
                        start_color=GREEN, end_color=GREEN, fill_type="solid"
                    )
                elif c.value == "로켓직구":
                    c.fill = PatternFill(
                        start_color=RED, end_color=RED, fill_type="solid"
                    )

            # 아이템 위너비율 100% 초록색 (어제 기준)
            if sub == col_위너비율_어제 and c.value not in (None, ""):
                try:
                    v = float(c.value)
                    if abs(v - 100.0) < 0.0001:
                        c.fill = PatternFill(
                            start_color=GREEN, end_color=GREEN, fill_type="solid"
                        )
                except Exception:
                    pass

            # 링크 처리
            if grp == "링크" and c.value and str(c.value).startswith("http"):
                c.hyperlink = str(c.value)
                c.value = "Link"
                c.font = Font(color="0563C1", underline="single")
                c.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{ws.max_row}"

    wb.save(path)


def main():
    from config.settings import Config

    df, curr_date, prev_date = get_data(Config.INTEGRATED_DB_PATH)

    if not df.empty and curr_date and prev_date:
        Config.OUTPUT_DIR.mkdir(exist_ok=True)
        output = (
            Config.OUTPUT_DIR
            / f"price_agent_dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        create_excel(df, str(output), curr_date, prev_date)
        print(f"✅ 완료: {output}")
    else:
        print("❌ 데이터 없음 또는 스냅샷 부족")


if __name__ == "__main__":
    main()
