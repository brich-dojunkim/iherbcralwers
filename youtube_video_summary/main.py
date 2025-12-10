"""
YouTube 채널 동영상 정보 수집 및 요약 스크립트
Groq Whisper API(초고속 STT) + Gemini(텍스트 요약)
"""

import json
import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

import time
from datetime import datetime, timedelta
from typing import List, Dict

import yt_dlp
from google import genai
from google.genai.types import HttpOptions
from groq import Groq  # Groq API 추가

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from faster_whisper import WhisperModel
    WHISPER_AVAILABLE = True
except ImportError:
    WHISPER_AVAILABLE = False
    print("⚠️  faster-whisper 미설치. Groq만 사용 가능.")
    print("   설치: pip install faster-whisper")

# YouTube 채널 목록
CHANNELS = [
    "https://www.youtube.com/@yakstory119",
    "https://www.youtube.com/@Ojingeryaksa",
    "https://www.youtube.com/@리틀약사",
    "https://www.youtube.com/@약사메디슨맨",
    "https://www.youtube.com/@양과자",
    "https://www.youtube.com/@HongSee_yaksa"
]


# ─────────────────────────────────────────────────────
# Rate Limit 추적 클래스
# ─────────────────────────────────────────────────────

class RateLimitTracker:
    """
    Groq Whisper API의 ASH (Audio Seconds per Hour) 제한 추적
    슬라이딩 윈도우 방식으로 과거 1시간 동안의 오디오 처리량 추적
    """
    def __init__(self, limit_ash: int = 7000):  # 7200초 중 200초는 버퍼
        self.limit_ash = limit_ash
        self.requests = []  # [(timestamp, duration), ...]
    
    def _clean_old_requests(self):
        """1시간 이상 된 요청 제거"""
        now = time.time()
        self.requests = [(t, d) for t, d in self.requests if now - t < 3600]
    
    def get_current_ash(self) -> int:
        """현재 1시간 윈도우의 누적 오디오 초"""
        self._clean_old_requests()
        return sum(d for _, d in self.requests)
    
    def can_process(self, duration: int) -> bool:
        """지정된 duration을 처리할 수 있는지 확인"""
        current = self.get_current_ash()
        return current + duration <= self.limit_ash
    
    def wait_time_needed(self, duration: int) -> float:
        """
        duration을 처리하기 위해 필요한 대기 시간 (초)
        가장 오래된 요청이 윈도우에서 빠질 때까지 대기
        """
        self._clean_old_requests()
        current = self.get_current_ash()
        
        if current + duration <= self.limit_ash:
            return 0
        
        # 목표: current_ash가 limit - duration 이하로 떨어질 때까지
        target = self.limit_ash - duration
        
        if not self.requests:
            return 0
        
        # 요청들을 시간순으로 제거하면서 계산
        accumulated = current
        now = time.time()
        
        for timestamp, req_duration in sorted(self.requests, key=lambda x: x[0]):
            accumulated -= req_duration
            if accumulated <= target:
                # 이 요청이 만료될 때까지 대기
                wait = (timestamp + 3600) - now
                return max(wait, 0) + 10  # 10초 버퍼
        
        # 모든 요청이 만료되어야 함
        oldest_time = self.requests[0][0]
        return max((oldest_time + 3600) - now, 0) + 10
    
    def record(self, duration: int):
        """처리한 오디오 duration 기록"""
        self.requests.append((time.time(), duration))
    
    def get_stats(self) -> Dict:
        """현재 상태 통계"""
        current = self.get_current_ash()
        return {
            'current_ash': current,
            'limit_ash': self.limit_ash,
            'remaining_ash': self.limit_ash - current,
            'usage_percent': (current / self.limit_ash) * 100,
            'requests_in_window': len(self.requests)
        }


# 중간 저장 파일 경로
CHECKPOINT_DIR = "checkpoints"
VIDEOS_CHECKPOINT = os.path.join(CHECKPOINT_DIR, "collected_videos.json")
SUMMARIES_CHECKPOINT = os.path.join(CHECKPOINT_DIR, "summaries.json")
TRANSCRIPTS_CHECKPOINT = os.path.join(CHECKPOINT_DIR, "transcripts.csv")

# 오디오 저장 디렉토리
AUDIO_DIR = "audio"

# Whisper 모델 캐시 (로컬 폴백용)
_WHISPER_MODEL = None


def get_whisper_model(model_size: str = "small"):
    """Whisper 로컬 모델 로드 (폴백용)"""
    global _WHISPER_MODEL
    if _WHISPER_MODEL is None:
        if not WHISPER_AVAILABLE:
            raise Exception("faster-whisper가 설치되지 않았습니다")
        print(f"   🧠 Whisper 로컬 모델 로딩... (size={model_size})")
        _WHISPER_MODEL = WhisperModel(model_size, device="cpu", compute_type="int8")
        print(f"   ✅ 모델 로딩 완료")
    return _WHISPER_MODEL


def transcribe_with_local_whisper(audio_path: str) -> str:
    """
    로컬 Whisper STT (Groq 폴백용)
    느리지만 제한 없음
    """
    print(f"   🔄 로컬 Whisper STT 사용 (Groq 폴백)")
    
    model = get_whisper_model(model_size="small")
    
    file_size_mb = os.path.getsize(audio_path) / (1024 * 1024)
    print(f"   📦 파일 크기: {file_size_mb:.1f}MB")
    
    start_time = time.time()
    segments, info = model.transcribe(
        audio_path,
        language="ko",
        beam_size=5
    )
    
    print(f"   ⏳ 음성 길이: {info.duration:.0f}초")
    
    texts = []
    for seg in segments:
        texts.append(seg.text.strip())
    
    transcript = " ".join(texts)
    elapsed = time.time() - start_time
    
    print(f"   ✅ 로컬 STT 완료 ({len(transcript)}자, {elapsed:.1f}초 소요)")
    
    # 오디오 파일 삭제
    try:
        os.remove(audio_path)
        print(f"   🗑️  오디오 파일 삭제")
    except Exception as e:
        print(f"   ⚠️  파일 삭제 실패: {e}")
    
    return transcript


def ensure_checkpoint_dir():
    if not os.path.exists(CHECKPOINT_DIR):
        os.makedirs(CHECKPOINT_DIR)


def ensure_audio_dir():
    if not os.path.exists(AUDIO_DIR):
        os.makedirs(AUDIO_DIR)


def save_videos_checkpoint(videos: List[Dict]):
    ensure_checkpoint_dir()
    with open(VIDEOS_CHECKPOINT, 'w', encoding='utf-8') as f:
        json.dump(videos, f, ensure_ascii=False, indent=2)
    print(f"   💾 동영상 정보 저장: {len(videos)}개")


def load_videos_checkpoint() -> List[Dict]:
    if os.path.exists(VIDEOS_CHECKPOINT):
        with open(VIDEOS_CHECKPOINT, 'r', encoding='utf-8') as f:
            videos = json.load(f)
        print(f"   📂 기존 동영상 로드: {len(videos)}개")
        return videos
    return []


def save_summaries_checkpoint(summaries: List[Dict]):
    ensure_checkpoint_dir()
    with open(SUMMARIES_CHECKPOINT, 'w', encoding='utf-8') as f:
        json.dump(summaries, f, ensure_ascii=False, indent=2)
    print(f"   💾 요약 결과 저장: {len(summaries)}개")


def load_summaries_checkpoint() -> List[Dict]:
    if os.path.exists(SUMMARIES_CHECKPOINT):
        with open(SUMMARIES_CHECKPOINT, 'r', encoding='utf-8') as f:
            summaries = json.load(f)
        print(f"   📂 기존 요약 로드: {len(summaries)}개")
        return summaries
    return []


def save_transcripts_checkpoint(transcripts: Dict[str, str]):
    """transcript를 CSV로 저장"""
    ensure_checkpoint_dir()
    data = [{'video_id': vid, 'transcript': text, 'length': len(text)} 
            for vid, text in transcripts.items()]
    df = pd.DataFrame(data)
    df.to_csv(TRANSCRIPTS_CHECKPOINT, index=False, encoding='utf-8-sig')
    print(f"   💾 transcript 저장: {len(transcripts)}개")


def load_transcripts_checkpoint() -> Dict[str, str]:
    """저장된 transcript 불러오기"""
    if os.path.exists(TRANSCRIPTS_CHECKPOINT):
        df = pd.read_csv(TRANSCRIPTS_CHECKPOINT, encoding='utf-8-sig')
        transcripts = dict(zip(df['video_id'], df['transcript']))
        print(f"   📂 기존 transcript 로드: {len(transcripts)}개")
        return transcripts
    return {}


def clear_checkpoints():
    files_to_remove = [VIDEOS_CHECKPOINT, SUMMARIES_CHECKPOINT, TRANSCRIPTS_CHECKPOINT]
    for filepath in files_to_remove:
        if os.path.exists(filepath):
            os.remove(filepath)
    print(f"   🗑️  체크포인트 정리 완료")


def get_channel_videos(channel_url: str, days_back: int = 30) -> List[Dict]:
    """채널의 최근 동영상 정보 수집"""
    cutoff_date = datetime.now() - timedelta(days=days_back)
    
    ydl_opts = {
        'quiet': True,
        'no_warnings': True,
        'extract_flat': False,
        'playlistend': 50,
        'skip_download': True,
        'ignoreerrors': True,
        'noprogress': True,
    }
    
    videos = []
    
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            print(f"   ⏳ 채널 정보 로딩 중...")
            result = ydl.extract_info(f"{channel_url}/videos", download=False)
            
            if 'entries' in result:
                channel_name = result.get('channel', 'Unknown')
                print(f"   ✅ 채널: {channel_name}")
                
                entries = [e for e in result['entries'] if e]
                collected = 0
                
                for entry in entries:
                    try:
                        if entry.get('availability') == 'subscriber_only':
                            continue
                        
                        upload_date = datetime.strptime(entry.get('upload_date', '20000101'), '%Y%m%d')
                        
                        if upload_date >= cutoff_date:
                            video_id = entry.get('id', '')
                            videos.append({
                                'channel': channel_name,
                                'channel_url': channel_url,
                                'title': entry.get('title', ''),
                                'video_id': video_id,
                                'url': f"https://www.youtube.com/watch?v={video_id}",
                                'upload_date': entry.get('upload_date', ''),
                                'view_count': entry.get('view_count', 0),
                                'duration': entry.get('duration', 0),
                                'description': entry.get('description', '')[:500],
                            })
                            collected += 1
                    except:
                        continue
                
                print(f"   ✅ 수집: {collected}개")
    
    except Exception as e:
        print(f"   ❌ 오류: {str(e)[:100]}")
    
    return videos


def download_audio_for_video(video_info: Dict) -> str:
    """오디오만 다운로드"""
    ensure_audio_dir()
    video_url = video_info["url"]
    video_id = video_info["video_id"]
    
    ydl_opts = {
        "quiet": True,
        "no_warnings": True,
        "format": "bestaudio/best",
        "outtmpl": os.path.join(AUDIO_DIR, f"{video_id}.%(ext)s"),
        "noprogress": True,  # 진행바 숨기기
    }
    
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        print(f"   🎧 오디오 다운로드 중...")
        info = ydl.extract_info(video_url, download=True)
        filename = ydl.prepare_filename(info)
    
    return filename


def transcribe_with_groq(audio_path: str) -> str:
    """
    Groq Whisper API로 초고속 STT
    - 파일 크기 제한: 25MB
    - Rate Limit 발생 시 retry-after 기반 대기
    """
    print(f"   📝 Groq Whisper API 호출 중...")
    
    groq_api_key = os.getenv('GROQ_API_KEY')
    if not groq_api_key:
        raise Exception("GROQ_API_KEY 환경변수가 설정되지 않았습니다")
    
    # 파일 크기 확인
    file_size_mb = os.path.getsize(audio_path) / (1024 * 1024)
    print(f"   📦 파일 크기: {file_size_mb:.1f}MB")
    
    # 25MB 초과 시 에러
    if file_size_mb > 25:
        raise Exception(f"파일 크기 초과 ({file_size_mb:.1f}MB > 25MB). 영상이 너무 깁니다.")
    
    client = Groq(api_key=groq_api_key)
    
    # Rate Limit 대응: 최대 3회 재시도
    max_retries = 3
    for attempt in range(max_retries):
        try:
            with open(audio_path, "rb") as audio_file:
                start_time = time.time()
                transcription = client.audio.transcriptions.create(
                    file=audio_file,
                    model="whisper-large-v3-turbo",
                    language="ko",
                    response_format="text"
                )
                elapsed = time.time() - start_time
            
            transcript = transcription
            print(f"   ✅ STT 완료 ({len(transcript)}자, {elapsed:.1f}초 소요)")
            
            # 오디오 파일 삭제
            try:
                os.remove(audio_path)
                print(f"   🗑️  오디오 파일 삭제")
            except Exception as e:
                print(f"   ⚠️  파일 삭제 실패: {e}")
            
            return transcript
            
        except Exception as e:
            error_str = str(e)
            
            # Rate Limit 에러인 경우
            if "429" in error_str or "rate limit" in error_str.lower():
                # 에러 메시지 전체 출력 (디버깅용)
                print(f"   ⚠️  Rate Limit 발생")
                
                # retry 시간 파싱
                wait_time = parse_retry_time(error_str)
                
                if attempt < max_retries - 1:
                    print(f"   ⏳ {wait_time}초 ({wait_time/60:.1f}분) 대기 후 재시도... ({attempt+1}/{max_retries})")
                    time.sleep(wait_time)
                    continue
                else:
                    raise Exception(f"Rate Limit 초과 ({max_retries}회 재시도 실패). 나중에 다시 시도하세요.")
            else:
                # 다른 에러는 바로 raise
                raise


def parse_retry_time(error_msg: str) -> int:
    """
    Groq 에러 메시지에서 정확한 대기 시간 파싱
    실제 형식: "Please try again in 7m36.5s"
    """
    import re
    
    # 모든 시간 패턴 찾기 (분과 초 조합)
    # 7m36.5s, 52m3s, 7m36s 등 모두 매칭
    pattern = r'(\d+)m([\d.]+)s'
    match = re.search(pattern, error_msg)
    
    if match:
        minutes = int(match.group(1))
        seconds = float(match.group(2))
        total = minutes * 60 + int(seconds) + 10  # 10초 버퍼
        print(f"   🔍 대기 시간 파싱: {minutes}분 {seconds:.1f}초 → {total}초 대기")
        return total
    
    # 초만 있는 경우: "45s"
    pattern2 = r'(\d+)s'
    match = re.search(pattern2, error_msg)
    if match:
        seconds = int(match.group(1)) + 10
        print(f"   🔍 대기 시간 파싱: {seconds-10}초 → {seconds}초 대기")
        return seconds
    
    # 파싱 실패 (발생하면 안 됨)
    print(f"   ⚠️  파싱 실패! 기본 10분 대기")
    print(f"   📄 에러 메시지: {error_msg[:200]}")
    return 600


def summarize_video_with_gemini(video_info: Dict, transcript: str) -> str:
    """Gemini로 텍스트 요약"""
    api_key = os.getenv('GEMINI_API_KEY')
    if not api_key:
        return "요약 실패: GEMINI_API_KEY 미설정"
    
    client = genai.Client(
        api_key=api_key,
        http_options=HttpOptions(api_version="v1")
    )
    
    prompt = f"""약사 유튜브 영상 분석 - 건강기능식품 정보 추출

제목: {video_info['title']}
채널: {video_info['channel']}

[받아쓰기 내용]
{transcript}

다음 형식으로 요약:

**1. 핵심 주제** (1문장)

**2. 제품/성분 정보**
📦 제품명: [브랜드+제품명] 또는 "언급 없음"
🧪 성분명: [구체적 성분] 또는 "언급 없음"  
💊 용도: [건강 목적]
👨‍⚕️ 약사 의견: [추천/주의/중립]

**3. 주요 내용** (3-5개 포인트)

**4. 트렌드 시사점** (1-2문장)
"""
    
    try:
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt
        )
        return response.text
    except Exception as e:
        if "429" in str(e) or "quota" in str(e).lower():
            return "요약 실패: 쿼터 초과"
        return f"요약 실패: {str(e)[:200]}"


def save_to_excel(all_videos: List[Dict], summaries: List[Dict], output_filename: str):
    """엑셀 저장"""
    print(f"\n📊 엑셀 생성 중...")
    
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # 1. 전체 동영상
        df_all = pd.DataFrame(all_videos)
        df_all['업로드일'] = pd.to_datetime(df_all['upload_date'], format='%Y%m%d').dt.strftime('%Y-%m-%d')
        
        df_display = df_all[['channel', 'title', '업로드일', 'view_count', 'duration', 'url']].copy()
        df_display.columns = ['채널', '제목', '업로드일', '조회수', '길이(초)', 'URL']
        df_display['조회수'] = df_display['조회수'].apply(lambda x: f"{x:,}")
        df_display.to_excel(writer, sheet_name='전체 동영상 목록', index=False)
        
        # 2. AI 요약
        if summaries:
            summary_data = []
            for item in summaries:
                v = item['video_info']
                summary_data.append({
                    '채널': v['channel'],
                    '제목': v['title'],
                    '업로드일': datetime.strptime(v['upload_date'], '%Y%m%d').strftime('%Y-%m-%d'),
                    '조회수': f"{v['view_count']:,}",
                    'URL': v['url'],
                    'AI 요약': item['summary']
                })
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='AI 요약 결과', index=False)
        
        # 3. 채널 통계
        stats = df_all.groupby('channel').agg({
            'title': 'count',
            'view_count': ['sum', 'mean', 'max'],
            'duration': 'sum'
        }).reset_index()
        stats.columns = ['채널', '동영상 수', '총 조회수', '평균 조회수', '최대 조회수', '총 길이(초)']
        for col in ['총 조회수', '평균 조회수', '최대 조회수']:
            stats[col] = stats[col].apply(lambda x: f"{int(x):,}")
        stats['총 길이(분)'] = (stats['총 길이(초)'].astype(int) / 60).round(1)
        stats = stats.drop('총 길이(초)', axis=1)
        stats.to_excel(writer, sheet_name='채널별 통계', index=False)
    
    print(f"   ✅ 엑셀 생성 완료")


def main():
    start_time = datetime.now().timestamp()
    
    print("=" * 80)
    print("🎬 YouTube 약사 채널 분석 시스템 (Groq Whisper + Gemini)")
    print("=" * 80)
    
    days_back = 30
    print(f"\n📅 수집 기간: 최근 {days_back}일")
    print(f"📺 대상 채널: {len(CHANNELS)}개")
    
    # 체크포인트 확인
    existing_videos = load_videos_checkpoint()
    existing_summaries = load_summaries_checkpoint()
    existing_transcripts = load_transcripts_checkpoint()
    
    if existing_videos or existing_summaries:
        print(f"\n⚠️  이전 작업 발견")
        print(f"   동영상: {len(existing_videos)}개")
        print(f"   요약: {len(existing_summaries)}개")
        print(f"   transcript: {len(existing_transcripts)}개")
        response = input("이어서 진행? (y/n): ").lower()
        if response == 'y':
            all_videos = existing_videos
            summaries = existing_summaries
            transcripts = existing_transcripts
            skip_collection = True
        else:
            clear_checkpoints()
            all_videos = []
            summaries = []
            transcripts = {}
            skip_collection = False
    else:
        all_videos = []
        summaries = []
        transcripts = {}
        skip_collection = False
    
    # 1단계: 동영상 수집
    if not skip_collection:
        print("\n" + "=" * 80)
        print("🔍 1단계: 동영상 정보 수집")
        print("=" * 80)
        
        for idx, channel_url in enumerate(CHANNELS, 1):
            print(f"\n[{idx}/{len(CHANNELS)}] {channel_url.split('@')[1]}")
            videos = get_channel_videos(channel_url, days_back)
            all_videos.extend(videos)
            save_videos_checkpoint(all_videos)
        
        print(f"\n✅ 수집 완료: {len(all_videos)}개")
        all_videos.sort(key=lambda x: x['upload_date'], reverse=True)
    
    # 2단계: STT + 요약
    print("\n" + "=" * 80)
    print("🤖 2단계: STT + AI 요약")
    print("=" * 80)
    
    summarized_ids = {s['video_info']['video_id'] for s in summaries}
    videos_to_summarize = [v for v in all_videos if v['video_id'] not in summarized_ids]
    
    print(f"\n완료: {len(summaries)}, 대기: {len(videos_to_summarize)}")
    
    if videos_to_summarize:
        # 총 오디오 길이 계산
        total_audio_seconds = sum(v['duration'] for v in videos_to_summarize)
        estimated_hours = total_audio_seconds / 3600
        
        print(f"⏱️  총 오디오 길이: {total_audio_seconds:,}초 ({estimated_hours:.1f}시간)")
        print(f"🔧 Groq Whisper API + Gemini")
        print(f"📊 ASH 제한: 시간당 7,200초")
        
        # Rate Limit Tracker 초기화
        rate_tracker = RateLimitTracker(limit_ash=7000)
        
        for i, video in enumerate(videos_to_summarize, 1):
            video_id = video['video_id']
            duration = video['duration']
            
            # all_videos에서의 실제 인덱스 찾기
            actual_index = next((idx for idx, v in enumerate(all_videos, 1) if v['video_id'] == video_id), i)
            
            print(f"\n[{actual_index}/{len(all_videos)}] {video['title'][:50]}")
            print(f"   ⏱️  영상 길이: {duration}초 ({duration/60:.1f}분)")
            
            # Rate Limit 체크 및 대기
            if not rate_tracker.can_process(duration):
                wait_time = rate_tracker.wait_time_needed(duration)
                stats = rate_tracker.get_stats()
                
                print(f"   ⚠️  ASH 제한 근접 ({stats['current_ash']}/{stats['limit_ash']}초)")
                print(f"   ⏳ {wait_time}초 ({wait_time/60:.1f}분) 대기 중...")
                
                time.sleep(wait_time)
                print(f"   ✅ 대기 완료")
            
            try:
                # STT: 기존 transcript 있으면 재사용
                if video_id in transcripts:
                    print(f"   ♻️  기존 transcript 재사용")
                    transcript = transcripts[video_id]
                else:
                    audio_path = download_audio_for_video(video)
                    transcript = transcribe_with_groq(audio_path)
                    transcripts[video_id] = transcript
                    save_transcripts_checkpoint(transcripts)
                    
                    # STT 성공 시 duration 기록
                    rate_tracker.record(duration)
                
                # 요약
                summary = summarize_video_with_gemini(video, transcript)
                
            except Exception as e:
                print(f"   ❌ 오류: {str(e)[:200]}")
                summary = f"처리 실패: {str(e)[:200]}"
            
            summaries.append({
                'video_info': video,
                'summary': summary,
                'processed_at': datetime.now().isoformat()
            })
            
            save_summaries_checkpoint(summaries)
            
            # 진행 상황 표시
            stats = rate_tracker.get_stats()
            print(f"   📊 ASH: {stats['current_ash']}/{stats['limit_ash']}초 ({stats['usage_percent']:.1f}%)")
            
            # 영상 간 짧은 대기
            if i < len(videos_to_summarize):
                time.sleep(2)
    
    # 결과 저장
    excel_file = f'youtube_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    save_to_excel(all_videos, summaries, excel_file)
    
    print(f"\n✅ 완료: {excel_file}")
    print(f"⏱️  소요 시간: {(datetime.now().timestamp() - start_time) / 60:.1f}분")
    
    clear_checkpoints()
    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()